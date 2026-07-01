"""GP project helpers: template resolution, save, Excel export."""
import io
import re
import uuid

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from .models import (
    FacultySubjectAssignment, FormField, FormTemplate,
    GPGroup, GPGroupMemberDetail, ProjectCase, Subject,
)

MAX_GP_GROUP_MEMBERS = 3


YES_NO_CHOICES = [('Yes', 'Yes'), ('No', 'No')]
GENDER_CHOICES = [('Male', 'Male'), ('Female', 'Female'), ('Other', 'Other')]
RELIGION_CHOICES = [
    ('Hindu', 'Hindu'),
    ('Muslim', 'Muslim'),
    ('Christian', 'Christian'),
    ('Sikh', 'Sikh'),
    ('Jain', 'Jain'),
    ('Buddhist', 'Buddhist'),
    ('Other', 'Other'),
]


def is_gp_submission_locked(dept):
    """True when the department GP submission deadline has passed."""
    if not dept or not dept.gp_submission_deadline:
        return False
    return timezone.now() >= dept.gp_submission_deadline


def slugify_field_name(label):
    name = re.sub(r'[^a-zA-Z0-9]+', '_', label.lower()).strip('_')
    return name or 'field'


def normalize_label(label):
    if label is None:
        return ''
    return re.sub(r'\s+', ' ', str(label).replace('\n', ' ')).strip()


# Columns already built into the student GP form — skip when importing from Excel.
FIXED_FIELD_LABELS = {
    'case', 'project title', 'gender diversity', 'religion diversity',
    'enrollment no', 'name of the student', 'name', 'sem iii roll no',
    'roll no', 'div', 'gender', 'gender (m/f)', 'gender m/f',
    'gender (m/f)', 'gender diversity (y/n)', 'religion diversity (y/n)',
    'area name', 'religion',
}

HEADER_KEYWORDS = ('enrollment', 'case', 'project title', 'roll', 'div', 'gender', 'group id')


def _is_fixed_field(label):
    low = normalize_label(label).lower()
    if not low or low == 'none':
        return True
    if low in FIXED_FIELD_LABELS:
        return True
    for fixed in FIXED_FIELD_LABELS:
        if low.startswith(fixed) or fixed in low:
            if 'diversity' in low or low in ('case', 'project title', 'enrollment no', 'div'):
                return True
            if 'roll' in low and 'group' not in low:
                return True
            if low.startswith('gender') and 'diversity' not in low:
                return True
            if 'name of the student' in low or low == 'name':
                return True
    return False


def _detect_field_type(label):
    low = normalize_label(label).lower()
    if 'y/n' in low or 'diversity' in low:
        return FormField.FieldType.YES_NO
    if 'm/f' in low:
        return FormField.FieldType.SELECT
    return FormField.FieldType.TEXT


def _detect_field_scope(label):
    low = normalize_label(label).lower()
    member_hints = ('enrollment', 'name of the student', 'roll', 'div', 'gender')
    if any(h in low for h in member_hints) and 'diversity' not in low and 'group id' not in low:
        return FormField.FieldScope.MEMBER
    return FormField.FieldScope.GROUP


def find_excel_header_row(file_path_or_file):
    """Locate the real header row in a GP project details Excel file."""
    wb = openpyxl.load_workbook(file_path_or_file, read_only=True, data_only=True)
    ws = wb.active
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx > 30:
            break
        cells = [normalize_label(c) for c in row]
        non_empty = [c for c in cells if c]
        if len(non_empty) < 3:
            continue
        low_cells = [c.lower() for c in non_empty]
        matches = sum(1 for c in low_cells if any(k in c for k in HEADER_KEYWORDS))
        if matches >= 3:
            wb.close()
            return row_idx, cells
    wb.close()
    return None, []


def parse_gp_excel_fields(file):
    """
    Parse GP project Excel and return field definitions.
    Skips title/note rows and columns already on the student form.
    """
    header_row_idx, headers = find_excel_header_row(file)
    if not header_row_idx:
        return []

    fields = []
    seen = set()
    for col_idx, raw in enumerate(headers):
        label = normalize_label(raw)
        if not label:
            continue
        low = label.lower()
        if _is_fixed_field(label):
            continue
        if low in seen:
            continue
        seen.add(low)

        ftype = _detect_field_type(label)
        choices = ''
        if ftype == FormField.FieldType.SELECT:
            choices = 'M, F' if 'm/f' in low else ''

        fields.append({
            'field_label': label,
            'field_name': slugify_field_name(label),
            'field_scope': _detect_field_scope(label),
            'field_type': ftype,
            'choices': choices,
            'order': col_idx,
            'is_required': False,
        })
    return fields


def import_fields_from_excel(file, template, replace=True):
    """Import dynamic form fields from GP project Excel template."""
    parsed = parse_gp_excel_fields(file)
    if replace:
        FormField.objects.filter(template=template).delete()

    created = 0
    for item in parsed:
        if not replace and FormField.objects.filter(template=template, field_label=item['field_label']).exists():
            continue
        FormField.objects.create(template=template, **item)
        created += 1
    return created, parsed


def import_cases_from_excel(file, semester, department=None):
    """Extract CASE values from Excel and create ProjectCase entries."""
    header_row_idx, headers = find_excel_header_row(file)
    if not header_row_idx:
        return 0

    case_col = None
    for i, h in enumerate(headers):
        if normalize_label(h).lower() == 'case':
            case_col = i
            break
    if case_col is None:
        return 0

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    cases_found = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= header_row_idx:
            continue
        if case_col >= len(row):
            continue
        val = normalize_label(row[case_col])
        if val and val.upper().startswith('CASE'):
            cases_found.add(val)

    created = 0
    for name in sorted(cases_found):
        _, was_created = ProjectCase.objects.get_or_create(
            name=name,
            semester=semester,
            department=department,
            defaults={'is_active': True},
        )
        if was_created:
            created += 1
    wb.close()
    return created


def get_gp_template(department):
    if not department:
        return None
    return FormTemplate.objects.filter(
        exam_type='GP',
        semester=department.semester,
    ).filter(
        Q(department=department) | Q(department__isnull=True)
    ).order_by('-department_id', '-created_at').first()


def get_project_cases(department):
    if not department:
        return ProjectCase.objects.none()
    return ProjectCase.objects.filter(
        is_active=True,
        semester=department.semester,
    ).filter(
        Q(department=department) | Q(department__isnull=True)
    ).order_by('name')


def get_group_fields(template):
    if not template:
        return []
    return list(template.fields.filter(field_scope=FormField.FieldScope.GROUP))


def get_member_fields(template):
    if not template:
        return []
    return list(template.fields.filter(field_scope=FormField.FieldScope.MEMBER))


def parse_group_data_from_post(post, template):
    data = {}
    for field in get_group_fields(template):
        val = post.get(f'group_{field.field_name}', '').strip()
        if val:
            data[field.field_name] = val
    return data


def get_faculty_initials(department, batch, subject):
    """Faculty mentor code for batch + subject from admin assignments."""
    if not department or not batch or not subject:
        return ''
    assignment = FacultySubjectAssignment.objects.filter(
        department=department,
        batch=batch,
        subject=subject,
        is_active=True,
    ).select_related('faculty').first()
    if not assignment:
        return ''
    fac = assignment.faculty
    return (fac.mentor_code or '').strip() or fac.name.split()[0][:6].upper()


def get_faculty_name_for_subject(department, batch, subject):
    """Assigned faculty full name for batch + subject."""
    if not department or not batch or not subject:
        return ''
    assignment = FacultySubjectAssignment.objects.filter(
        department=department,
        batch=batch,
        subject=subject,
        is_active=True,
    ).select_related('faculty').first()
    if not assignment:
        return ''
    return assignment.faculty.name


def get_faculty_map_for_subjects(department, batch, subjects):
    return {
        str(s.pk): get_faculty_initials(department, batch, s)
        for s in subjects
    }


def get_subject_member_ids(department, subject_id, batch=None, exclude_group_id=None):
    """Student PKs already in a GP team for this subject (same batch when provided)."""
    if not department or not subject_id:
        return set()
    qs = GPGroup.objects.filter(department=department, subject_id=subject_id)
    if batch:
        qs = qs.filter(leader__batch=batch)
    if exclude_group_id:
        qs = qs.exclude(pk=exclude_group_id)
    return set(qs.values_list('members__pk', flat=True))


def get_taken_member_ids_any_group(department, batch, exclude_group_ids=None):
    """Student PKs already in any GP group for this department + batch."""
    if not department or not batch:
        return set()
    qs = GPGroup.objects.filter(department=department, leader__batch=batch)
    exclude_ids = set(exclude_group_ids or [])
    if exclude_ids:
        qs = qs.exclude(pk__in=exclude_ids)
    return set(qs.values_list('members__pk', flat=True))


def get_taken_members_by_subject(department, subjects, batch=None, exclude_group_id=None, exclude_group_ids=None):
    exclude_ids = set(exclude_group_ids or [])
    if exclude_group_id:
        exclude_ids.add(exclude_group_id)
    result = {}
    for s in subjects:
        taken = set(get_subject_member_ids(department, s.pk, batch=batch))
        if exclude_ids:
            for g in GPGroup.objects.filter(pk__in=exclude_ids, subject=s):
                taken -= set(g.members.values_list('pk', flat=True))
        result[str(s.pk)] = list(taken)
    return result


def get_taken_titles_by_subject(department, batch, subjects, exclude_group_ids=None):
    """Project titles already used in this batch, grouped by subject."""
    exclude_ids = set(exclude_group_ids or [])
    result = {str(s.pk): [] for s in subjects}
    if not department or not batch:
        return result

    qs = GPGroup.objects.filter(
        department=department,
        subject__in=subjects,
        leader__batch=batch,
    ).exclude(name='').select_related('subject')

    if exclude_ids:
        qs = qs.exclude(pk__in=exclude_ids)

    for group in qs.order_by('subject__name', 'name'):
        title = (group.name or '').strip()
        if not title:
            continue
        sid = str(group.subject_id)
        if sid in result and title not in result[sid]:
            result[sid].append(title)
    return result


def subjects_for_department(dept):
    if not dept:
        return Subject.objects.none()
    return Subject.objects.filter(
        Q(semester=dept.semester, department__isnull=True) | Q(department=dept)
    ).order_by('name')


def build_gp_subject_selection_options(subjects):
    """Theory subjects individually; all practical subjects as one combined option."""
    subject_list = list(subjects)
    theory = [s for s in subject_list if s.subject_type == Subject.SubjectType.THEORY]
    practical = [s for s in subject_list if s.subject_type == Subject.SubjectType.PRACTICAL]
    options = []
    for s in theory:
        options.append({
            'value': f'T-{s.pk}',
            'label': f'{s.name} (Theory)',
            'mode': 'theory',
            'subjects': [{'pk': s.pk, 'name': s.name}],
        })
    if practical:
        names = ' + '.join(s.name for s in practical)
        pks = ','.join(str(s.pk) for s in practical)
        options.append({
            'value': f'P-{pks}',
            'label': f'{names} (Practical)',
            'mode': 'practical',
            'subjects': [{'pk': s.pk, 'name': s.name} for s in practical],
        })
    return options


def parse_subject_selection(selection):
    """Return list of subject PKs from T-{id} or P-{id},{id} selection value."""
    if not selection:
        return []
    if selection.startswith('T-'):
        try:
            return [int(selection[2:])]
        except (TypeError, ValueError):
            return []
    if selection.startswith('P-'):
        ids = []
        for part in selection[2:].split(','):
            part = part.strip()
            if part.isdigit():
                ids.append(int(part))
        return ids
    if str(selection).isdigit():
        return [int(selection)]
    return []


def selection_value_for_groups(groups):
    """Build subject_selection value from one or more linked GP groups."""
    if not groups:
        return ''
    subjects = [g.subject for g in groups if g.subject_id]
    if not subjects:
        return ''
    if len(subjects) == 1 and subjects[0].subject_type == Subject.SubjectType.THEORY:
        return f'T-{subjects[0].pk}'
    pks = ','.join(str(s.pk) for s in subjects)
    return f'P-{pks}'


def get_bundle_groups(group):
    """All GP groups submitted together (practical bundle) or just this group."""
    if not group:
        return GPGroup.objects.none()
    if group.linked_batch_id:
        return GPGroup.objects.filter(linked_batch_id=group.linked_batch_id).select_related(
            'subject', 'project_case',
        ).order_by('subject__name')
    return GPGroup.objects.filter(pk=group.pk).select_related('subject', 'project_case')


def bundle_subject_entries_from_groups(groups):
    return [
        {
            'subject_id': g.subject_id,
            'subject_name': g.subject.name if g.subject_id else '',
            'case_id': g.project_case_id,
            'case_name': g.project_case.name if g.project_case_id else '',
            'title': g.name,
        }
        for g in groups
        if g.subject_id
    ]


def gp_case_uses_combined_title(case):
    """Case 1 — one shared project title for all subjects in a practical bundle."""
    if not case:
        return False
    name = normalize_label(case.name).lower()
    return bool(re.match(r'^case\s*1\b', name)) or name in ('case 1', 'case1')


def _resolve_subject_entry_from_post(post, sid, subj, case_bundle='', title_combined=''):
    """Build case_id and title for one subject from POST (bundle or per-subject fields)."""
    if case_bundle:
        case_id = case_bundle
        title = title_combined or post.get(f'title_{sid}', '').strip() or post.get('name', '').strip()
    else:
        case_id = post.get(f'case_{sid}', '').strip() or post.get('project_case', '').strip()
        title = post.get(f'title_{sid}', '').strip() or post.get('name', '').strip()
    return case_id, title


def get_faculty_initials_bundle(department, batch, subject_ids):
    parts = []
    for sid in subject_ids:
        subj = Subject.objects.filter(pk=sid).first()
        if not subj:
            continue
        initials = get_faculty_initials(department, batch, subj)
        if initials:
            parts.append(initials)
    return ' / '.join(parts)


def _apply_faculty_initials_single(group_data, template, department, batch, subject, force=False):
    """Auto-fill Subject Faculty from batch/subject assignment."""
    name = get_faculty_name_for_subject(department, batch, subject)
    initials = get_faculty_initials(department, batch, subject)
    display = name or initials
    if not display:
        return group_data
    data = dict(group_data or {})
    for key in ('subject_faculty_initials', 'faculty_initials', 'subject_faculty'):
        if force or not str(data.get(key, '')).strip():
            data[key] = display
    for field in get_group_fields(template):
        if 'faculty' in field.field_label.lower():
            if force or not str(data.get(field.field_name, '')).strip():
                data[field.field_name] = display
    return data


def apply_faculty_initials_to_group_data(group_data, template, department, batch, subject_or_ids):
    """Auto-fill faculty initials for one subject or a practical bundle."""
    if isinstance(subject_or_ids, (list, tuple)):
        subject_ids = list(subject_or_ids)
        if len(subject_ids) == 1:
            subj = Subject.objects.filter(pk=subject_ids[0]).first()
            return _apply_faculty_initials_single(group_data, template, department, batch, subj)
        initials = get_faculty_initials_bundle(department, batch, subject_ids)
        if not initials:
            return group_data
        data = dict(group_data or {})
        for key in ('subject_faculty_initials', 'faculty_initials', 'subject_faculty'):
            if key not in data or not str(data.get(key, '')).strip():
                data[key] = initials
        for field in get_group_fields(template):
            if 'faculty' in field.field_label.lower():
                if not str(data.get(field.field_name, '')).strip():
                    data[field.field_name] = initials
        return data
    return _apply_faculty_initials_single(group_data, template, department, batch, subject_or_ids)


def parse_member_data_from_post(post, student_pk, template):
    data = {}
    for field in get_member_fields(template):
        val = post.get(f'member_{student_pk}_{field.field_name}', '').strip()
        if val:
            data[field.field_name] = val
    return data


def save_gp_submission(student, dept, post, group=None):
    """Create or update GP group(s) from student POST data. Returns (primary_group, errors)."""
    from .models import Student

    errors = []
    if is_gp_submission_locked(dept):
        if group:
            errors.append('The GP project submission deadline has passed. You can no longer edit your project.')
        else:
            errors.append('The GP project submission deadline has passed. You can no longer submit a project.')
        return None, errors
    selection = post.get('subject_selection', '').strip()
    subject_ids = parse_subject_selection(selection)
    if not subject_ids and post.get('subject'):
        subject_ids = [int(post.get('subject'))]

    gender_div = post.get('gender_diversity', '')
    religion_div = post.get('religion_diversity', '')

    if not subject_ids:
        errors.append('Please select a subject.')
    if gender_div not in ('Yes', 'No'):
        errors.append('Please select Gender Diversity (Yes/No).')
    if religion_div not in ('Yes', 'No'):
        errors.append('Please select Religion Diversity (Yes/No).')

    member_ids = post.getlist('member_ids')
    if not member_ids:
        errors.append('Select at least one group member by enrollment number.')

    subject_entries = []
    case_bundle = post.get('case_bundle', '').strip()
    title_combined = post.get('title_combined', '').strip()
    bundle_case = ProjectCase.objects.filter(pk=case_bundle).first() if case_bundle else None
    combined_title_mode = gp_case_uses_combined_title(bundle_case)

    if len(subject_ids) > 1 and not case_bundle:
        errors.append('Please select a project case.')

    for sid in subject_ids:
        subj = Subject.objects.filter(pk=sid).first()
        if not subj:
            errors.append('Invalid subject selected.')
            continue
        case_id, title = _resolve_subject_entry_from_post(
            post, sid, subj, case_bundle=case_bundle, title_combined=title_combined,
        )
        if not case_id and len(subject_ids) == 1:
            errors.append(f'Please select a project case for {subj.name}.')
        if not title:
            if combined_title_mode:
                if not any('project title for all subjects' in e for e in errors):
                    errors.append('Please enter a project title for all subjects.')
            else:
                errors.append(f'Please enter a project title for {subj.name}.')
        subject_entries.append({
            'subject': subj,
            'case_id': case_id or None,
            'title': title,
        })

    if errors:
        return None, errors

    template = get_gp_template(dept)
    batch_students = Student.objects.filter(department=dept, batch=student.batch)
    valid_ids = set(batch_students.values_list('pk', flat=True))

    selected_pks = []
    for mid in member_ids:
        try:
            pk = int(mid)
        except (TypeError, ValueError):
            continue
        if pk in valid_ids:
            selected_pks.append(pk)

    if student.pk not in selected_pks:
        selected_pks.insert(0, student.pk)

    unique_members = list(dict.fromkeys(selected_pks))
    if len(unique_members) > MAX_GP_GROUP_MEMBERS:
        errors.append(f'A group can have at most {MAX_GP_GROUP_MEMBERS} members.')
        return None, errors
    selected_pks = unique_members

    exclude_ids = set()
    if group:
        bundle = list(get_bundle_groups(group))
        exclude_ids = {g.pk for g in bundle}

    taken_any = get_taken_member_ids_any_group(
        dept, student.batch, exclude_group_ids=list(exclude_ids),
    )
    for pk in selected_pks:
        if pk in taken_any:
            stu = batch_students.filter(pk=pk).first()
            label = stu.enrollment_no if stu else pk
            errors.append(
                f'{label} is already in another project group. '
                f'Remove them from that team first or pick a different student.'
            )

    if errors:
        return None, errors

    for pk in selected_pks:
        gender = post.get(f'member_gender_{pk}', '').strip()
        if not gender:
            stu = batch_students.filter(pk=pk).first()
            label = stu.enrollment_no if stu else pk
            errors.append(f'Please select gender for enrollment {label}.')

    if errors:
        return None, errors

    group_data = parse_group_data_from_post(post, template)

    batch_id = None
    existing_bundle = list(get_bundle_groups(group)) if group else []
    if len(subject_entries) > 1 or (existing_bundle and existing_bundle[0].linked_batch_id):
        batch_id = existing_bundle[0].linked_batch_id if existing_bundle else uuid.uuid4()

    saved_groups = []
    for idx, entry in enumerate(subject_entries):
        if existing_bundle:
            target = next((g for g in existing_bundle if g.subject_id == entry['subject'].pk), None)
            if target is None and idx < len(existing_bundle):
                target = existing_bundle[idx]
        else:
            target = None

        if target is None:
            target = GPGroup(leader=student, department=dept)

        target.subject = entry['subject']
        target.name = entry['title']
        target.project_case_id = entry['case_id']
        target.gender_diversity = gender_div
        target.religion_diversity = religion_div
        entry_group_data = _apply_faculty_initials_single(
            dict(group_data), template, dept, student.batch, entry['subject'], force=True,
        )
        target.group_data = entry_group_data
        target.linked_batch_id = batch_id
        target.is_submitted = True
        target.save()

        target.members.set(selected_pks)
        GPGroupMemberDetail.objects.filter(group=target).exclude(student_id__in=selected_pks).delete()

        for pk in selected_pks:
            detail, _ = GPGroupMemberDetail.objects.get_or_create(group=target, student_id=pk)
            detail.gender = post.get(f'member_gender_{pk}', '').strip()
            detail.region = post.get(f'member_region_{pk}', '').strip()
            detail.religion = post.get(f'member_religion_{pk}', '').strip()
            detail.member_data = parse_member_data_from_post(post, pk, template)
            detail.save()

        saved_groups.append(target)

    if existing_bundle:
        keep_ids = {g.pk for g in saved_groups}
        for old in existing_bundle:
            if old.pk not in keep_ids:
                old.delete()

    return saved_groups[0] if saved_groups else None, []


def _yn_export(val):
    if not val:
        return ''
    v = str(val).strip().lower()
    if v in ('yes', 'y'):
        return 'Y'
    if v in ('no', 'n'):
        return 'N'
    return val


def _gender_export(val):
    if not val:
        return ''
    v = str(val).strip().lower()
    if v in ('male', 'm'):
        return 'M'
    if v in ('female', 'f'):
        return 'F'
    return str(val)[:1].upper() if val else ''


def _read_group_id(group):
    """Return stored Group ID without generating a new one."""
    data = group.group_data or {}
    for key in ('group_id', 'groupid'):
        if data.get(key):
            return str(data[key])
    for field in get_group_fields(get_gp_template(group.department)):
        if 'group' in field.field_label.lower() and 'id' in field.field_label.lower():
            if data.get(field.field_name):
                return str(data[field.field_name])
    return ''


def _group_id_value(group, batch_counters):
    gid = _read_group_id(group)
    if gid:
        return gid
    batch = group.leader.batch or 'X'
    batch_counters[batch] = batch_counters.get(batch, 0) + 1
    return f"{batch}_{batch_counters[batch]}"


def _group_sort_key(group):
    """Sort by batch then Group ID number ascending (A1_1, A1_2, … A1_10)."""
    batch = group.leader.batch or 'Z'
    gid = _read_group_id(group)
    num = 0
    if gid and '_' in gid:
        try:
            num = int(gid.rsplit('_', 1)[-1])
        except ValueError:
            pass
    return (batch, num, gid)


# Alternating pair per batch division — groups alternate between the two shades.
BATCH_PALETTES = [
    ('D6EAF8', 'AED6F1'),  # A1 — blue
    ('D5F5E3', 'A9DFBF'),  # A2 — green
    ('FCF3CF', 'F9E79F'),  # A3 — yellow
    ('F5EEF8', 'D7BDE2'),  # A4 — purple
    ('FADBD8', 'F5B7B1'),  # A5 — red
    ('D1F2EB', 'A3E4D7'),  # A6 — teal
    ('FDEBD0', 'F5CBA7'),  # A7 — orange
    ('E8DAEF', 'C39BD3'),  # A8 — violet
    ('E5E8E8', 'BDC3C7'),  # A9 — grey
]


def _batch_palette(batch, all_batches):
    ordered = sorted(b for b in all_batches if b)
    try:
        idx = ordered.index(batch)
    except ValueError:
        idx = 0
    return BATCH_PALETTES[idx % len(BATCH_PALETTES)]


def _faculty_initials(group):
    batch = group.leader.batch if group.leader else ''
    if group.subject_id and group.department_id:
        name = get_faculty_name_for_subject(group.department, batch, group.subject)
        if name:
            return name
        initials = get_faculty_initials(group.department, batch, group.subject)
        if initials:
            return initials
    data = group.group_data or {}
    for key in ('subject_faculty_initials', 'faculty_initials', 'subject_faculty'):
        if data.get(key):
            return data[key]
    for field in get_group_fields(get_gp_template(group.department)):
        if 'faculty' in field.field_label.lower():
            return data.get(field.field_name, '')
    return ''


def get_gp_submitted_member_ids(department, subject_id, batch):
    """Student PKs already in a submitted GP group for batch + subject."""
    if not department or not subject_id or not batch:
        return set()
    return set(GPGroup.objects.filter(
        department=department,
        subject_id=subject_id,
        leader__batch=batch,
        is_submitted=True,
    ).values_list('members__pk', flat=True))


def get_pending_gp_students(department, subject_id, batch):
    """Batch students who have not submitted GP for this subject."""
    from .models import Student

    if not department or not subject_id or not batch:
        return Student.objects.none()
    submitted_ids = get_gp_submitted_member_ids(department, subject_id, batch)
    return Student.objects.filter(
        department=department,
        batch=batch,
    ).exclude(pk__in=submitted_ids).order_by('roll_no')


def export_gp_submissions_excel(groups, template=None, filename='gp_project_details.xlsx', subject=None):
    """
    Export GP submissions matching the official template layout:
    - Row 1: yellow merged title banner
    - Row 2: column headers
    - One row per student; Group ID & Project Title merged per group
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    groups = list(groups.select_related(
        'subject', 'project_case', 'leader', 'department', 'department__semester'
    ).prefetch_related('member_details__student', 'members'))
    groups.sort(key=_group_sort_key)

    if not subject and groups:
        subject = groups[0].subject

    extra_group_fields = [
        f for f in (get_group_fields(template) if template else [])
        if 'group id' not in f.field_label.lower() and 'faculty' not in f.field_label.lower()
    ]
    extra_member_fields = get_member_fields(template) if template else []

    std_headers = [
        'Group ID', 'CASE', 'Enrollment No', 'Name of the Student',
        'SEM III ROLL NO.', 'Div', 'Gender\n(M/F)',
        'Gender Diversity\n(Y/N)', 'Religion Diversity\n(Y/N)',
        'Project Title', 'Subject Faculty Name',
    ]
    tail_headers = ['Area Name', 'Religion']
    headers = (
        std_headers
        + [f.field_label for f in extra_group_fields]
        + [f.field_label for f in extra_member_fields]
        + tail_headers
    )
    num_cols = len(headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Project Details'

    thin = Side(style='thin', color='000000')
    thick = Side(style='medium', color='000000')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill('solid', fgColor='FFFF00')
    header_fill = PatternFill('solid', fgColor='B4C6E7')
    font_tnr = Font(name='Times New Roman', size=12)
    font_title = Font(name='Times New Roman', size=12, bold=True)
    font_header = Font(name='Times New Roman', size=12, bold=True)
    font_gid = Font(name='Times New Roman', size=12, bold=True, color='FF0000')

    # Title row
    dept = groups[0].department if groups else None
    sem_label = ''
    if dept:
        sem_label = dept.sheet_semester_label or (dept.semester.name if dept.semester else '')
    subj_name = subject.name if subject else (groups[0].subject.name if groups else 'GP')
    title_text = (
        'L.J.Institute of Engineering and Technology \n'
        f'Group Project Details - {subj_name}\n'
        f'{sem_label}'
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.fill = title_fill
    title_cell.font = font_title
    title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 55

    # Header row
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=label)
        cell.fill = header_fill
        cell.font = font_gid if col == 1 else font_header
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 36

    all_batches = {g.leader.batch for g in groups if g.leader and g.leader.batch}
    batch_group_idx = {}
    batch_counters = {}
    data_row = 3
    merge_ranges = []
    group_end_rows = []

    for group in groups:
        details = list(group.member_details.select_related('student').order_by('student__roll_no'))
        if not details:
            for m in group.members.order_by('roll_no'):
                details.append(GPGroupMemberDetail(student=m, gender='', member_data={}))

        if not details:
            continue

        gid = _group_id_value(group, batch_counters)
        case_name = group.project_case.name if group.project_case else ''
        title = group.name
        g_div = _yn_export(group.gender_diversity)
        r_div = _yn_export(group.religion_diversity)
        faculty = _faculty_initials(group)
        start_row = data_row
        batch = group.leader.batch or ''
        batch_group_idx[batch] = batch_group_idx.get(batch, 0) + 1
        light, dark = _batch_palette(batch, all_batches)
        group_fill = PatternFill(
            'solid',
            fgColor=dark if batch_group_idx[batch] % 2 == 0 else light,
        )

        for idx, detail in enumerate(details):
            s = detail.student
            gender = _gender_export(getattr(detail, 'gender', ''))
            member_data = getattr(detail, 'member_data', {}) or {}

            row_vals = [
                gid if idx == 0 else None,
                case_name,
                s.enrollment_no,
                s.name,
                s.roll_no,
                s.batch,
                gender,
                g_div,
                r_div,
                title if idx == 0 else None,
                faculty,
            ]
            for field in extra_group_fields:
                row_vals.append(group.group_data.get(field.field_name, '') if idx == 0 else '')
            for field in extra_member_fields:
                row_vals.append(member_data.get(field.field_name, ''))
            row_vals.append(getattr(detail, 'region', '') or '')
            row_vals.append(getattr(detail, 'religion', '') or '')

            is_last_in_group = idx == len(details) - 1
            row_border = Border(
                left=thin, right=thin, top=thin,
                bottom=thick if is_last_in_group else thin,
            )
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.font = font_tnr
                cell.fill = group_fill
                cell.border = row_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            data_row += 1

        end_row = data_row - 1
        if end_row >= start_row:
            merge_ranges.append((start_row, end_row, 1))   # Group ID
            merge_ranges.append((start_row, end_row, 10))  # Project Title
            merge_ranges.append((start_row, end_row, 11))  # Faculty Initials
            group_end_rows.append(end_row)

    for start_row, end_row, col in merge_ranges:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        merged = ws.cell(row=start_row, column=col)
        merged.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for end_row in group_end_rows:
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=end_row, column=col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)

    base_widths = [10, 10, 16, 28, 12, 6, 8, 12, 12, 30, 10]
    extra_count = max(0, num_cols - len(base_widths) - 2)
    widths = base_widths + [14] * extra_count + [18, 14]
    for i, w in enumerate(widths, 1):
        if i <= num_cols:
            ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def create_default_gp_template(semester, department=None, created_by=None):
    """Ensure a default GP template with standard + sample dynamic fields exists."""
    tmpl, created = FormTemplate.objects.get_or_create(
        name='GP Project Details',
        semester=semester,
        department=department,
        exam_type='GP',
        defaults={'created_by': created_by},
    )
    if created or tmpl.fields.count() == 0:
        FormField.objects.filter(template=tmpl).delete()
        # Standard fields are on GPGroup model; template holds extra dynamic fields only
    return tmpl


def departments_in_analytics_scope(user, ctx, dept=None):
    """Departments visible on GP analytics for the current admin."""
    from .models import Department, User

    if dept:
        return [dept]
    if user.role == User.Role.SEMESTER_ADMIN and ctx.get('semester'):
        return list(Department.objects.filter(semester=ctx['semester']).order_by('name'))
    if user.is_superuser or user.role == User.Role.SUPER_ADMIN:
        semester = ctx.get('semester')
        if semester:
            return list(Department.objects.filter(semester=semester).order_by('name'))
        return list(Department.objects.all().order_by('name'))
    assigned = ctx.get('department')
    return [assigned] if assigned else []


def compute_gp_analytics_for_departments(departments, batch=None):
    """Summary counts for GP project groups and students."""
    from .models import GPGroup, Student

    empty = {
        'total_students': 0,
        'submitted_students': 0,
        'pending_students': 0,
        'total_groups': 0,
        'submitted_groups': 0,
        'draft_groups': 0,
    }
    if not departments:
        return empty

    dept_ids = [d.pk for d in departments]
    students = Student.objects.filter(department_id__in=dept_ids)
    groups = GPGroup.objects.filter(department_id__in=dept_ids)
    submitted_groups_qs = GPGroup.objects.filter(department_id__in=dept_ids, is_submitted=True)

    if batch:
        students = students.filter(batch=batch)
        groups = groups.filter(leader__batch=batch)
        submitted_groups_qs = submitted_groups_qs.filter(leader__batch=batch)

    total_students = students.count()
    total_groups = groups.count()
    submitted_groups = groups.filter(is_submitted=True).count()

    student_ids = set(students.values_list('pk', flat=True))
    submitted_member_ids = set(submitted_groups_qs.values_list('members__pk', flat=True))
    submitted_students = len(student_ids & submitted_member_ids)
    pending_students = total_students - submitted_students

    return {
        'total_students': total_students,
        'submitted_students': submitted_students,
        'pending_students': pending_students,
        'total_groups': total_groups,
        'submitted_groups': submitted_groups,
        'draft_groups': total_groups - submitted_groups,
    }


def get_gp_analytics_batches(departments):
    from .models import Student

    if not departments:
        return []
    dept_ids = [d.pk for d in departments]
    return list(
        Student.objects.filter(department_id__in=dept_ids)
        .values_list('batch', flat=True)
        .distinct()
        .order_by('batch')
    )


def get_gp_analytics_student_rows(departments, batch=None, status_filter='all'):
    """Student list with GP submission status for analytics tables."""
    from .models import GPGroup, Student

    if not departments:
        return []

    dept_ids = [d.pk for d in departments]
    groups = GPGroup.objects.filter(
        department_id__in=dept_ids,
        is_submitted=True,
    ).select_related('subject', 'leader').prefetch_related('members')
    if batch:
        groups = groups.filter(leader__batch=batch)

    member_info = {}
    for group in groups:
        for member in group.members.all():
            if member.pk in member_info:
                continue
            member_info[member.pk] = {
                'group_title': group.name,
                'subject': group.subject.name if group.subject else '—',
                'is_leader': group.leader_id == member.pk,
            }

    students = Student.objects.filter(
        department_id__in=dept_ids,
    ).select_related('department').order_by('batch', 'roll_no')
    if batch:
        students = students.filter(batch=batch)

    rows = []
    for student in students:
        info = member_info.get(student.pk)
        submitted = info is not None
        if status_filter == 'submitted' and not submitted:
            continue
        if status_filter == 'pending' and submitted:
            continue
        rows.append({
            'student': student,
            'gp_status': 'Submitted' if submitted else 'Pending',
            'group_title': info['group_title'] if info else '—',
            'subject': info['subject'] if info else '—',
            'is_leader': info['is_leader'] if info else False,
        })
    return rows


def get_faculty_analytics_batches(faculty):
    from .models import FacultySubjectAssignment

    return list(
        FacultySubjectAssignment.objects.filter(faculty=faculty, is_active=True)
        .values_list('batch', flat=True)
        .distinct()
        .order_by('batch')
    )


def _faculty_assignment_group_query(faculty, batch=None):
    from .models import FacultySubjectAssignment

    assignments = FacultySubjectAssignment.objects.filter(faculty=faculty, is_active=True)
    if batch:
        assignments = assignments.filter(batch=batch)
    query = Q()
    for assignment in assignments:
        query |= Q(
            department_id=assignment.department_id,
            subject_id=assignment.subject_id,
            leader__batch=assignment.batch,
        )
    return query, assignments


def compute_gp_analytics_for_faculty(faculty, batch=None):
    """GP analytics limited to subjects and batches assigned to this faculty."""
    from .models import GPGroup, Student

    empty = {
        'total_students': 0,
        'submitted_students': 0,
        'pending_students': 0,
        'total_groups': 0,
        'submitted_groups': 0,
        'draft_groups': 0,
    }
    if not faculty or not faculty.department_id:
        return empty

    group_query, assignments = _faculty_assignment_group_query(faculty, batch=batch)
    if not assignments.exists():
        return empty

    assigned_batches = list(assignments.values_list('batch', flat=True).distinct())
    students = Student.objects.filter(
        department=faculty.department,
        batch__in=assigned_batches,
    )
    if batch:
        students = students.filter(batch=batch)

    groups = GPGroup.objects.filter(group_query) if group_query else GPGroup.objects.none()
    submitted_groups_qs = groups.filter(is_submitted=True)

    total_students = students.count()
    total_groups = groups.count()
    submitted_groups = submitted_groups_qs.count()

    student_ids = set(students.values_list('pk', flat=True))
    submitted_member_ids = set(submitted_groups_qs.values_list('members__pk', flat=True))
    submitted_students = len(student_ids & submitted_member_ids)
    pending_students = total_students - submitted_students

    return {
        'total_students': total_students,
        'submitted_students': submitted_students,
        'pending_students': pending_students,
        'total_groups': total_groups,
        'submitted_groups': submitted_groups,
        'draft_groups': total_groups - submitted_groups,
    }


def get_gp_analytics_student_rows_for_faculty(faculty, batch=None, status_filter='all'):
    from .models import GPGroup, Student

    if not faculty or not faculty.department_id:
        return []

    group_query, assignments = _faculty_assignment_group_query(faculty, batch=batch)
    if not assignments.exists():
        return []

    assigned_batches = list(assignments.values_list('batch', flat=True).distinct())
    groups = GPGroup.objects.filter(
        group_query,
        is_submitted=True,
    ).select_related('subject', 'leader').prefetch_related('members')
    if batch:
        groups = groups.filter(leader__batch=batch)

    member_info = {}
    for group in groups:
        for member in group.members.all():
            if member.pk in member_info:
                continue
            member_info[member.pk] = {
                'group_title': group.name,
                'subject': group.subject.name if group.subject else '—',
                'is_leader': group.leader_id == member.pk,
            }

    students = Student.objects.filter(
        department=faculty.department,
        batch__in=assigned_batches,
    ).order_by('batch', 'roll_no')
    if batch:
        students = students.filter(batch=batch)

    rows = []
    for student in students:
        info = member_info.get(student.pk)
        submitted = info is not None
        if status_filter == 'submitted' and not submitted:
            continue
        if status_filter == 'pending' and submitted:
            continue
        rows.append({
            'student': student,
            'gp_status': 'Submitted' if submitted else 'Pending',
            'group_title': info['group_title'] if info else '—',
            'subject': info['subject'] if info else '—',
            'is_leader': info['is_leader'] if info else False,
        })
    return rows
