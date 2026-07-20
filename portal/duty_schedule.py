"""Parse IPE duty schedule Excel and create faculty duty assignments.

Supports the flat template columns:
  subject | internal faculty | External Faculty Name | Batch | Duty Date | Time Slot | Email id

Also keeps backward compatibility with the older multi-block DIV/TIME schedule sheets.
"""
import re
from datetime import datetime, date

import openpyxl
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Q

from .models import (
    DutyScheduleUpload, Faculty, FacultyDutyAssignment, Student, Subject,
)


def department_batches(department):
    """Batches that belong to a department's student roster."""
    return list(
        Student.objects.filter(department=department)
        .values_list('batch', flat=True)
        .distinct()
        .order_by('batch')
    )


def _norm(val):
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%d-%m-%Y')
    if isinstance(val, date):
        return val.strftime('%d-%m-%Y')
    return re.sub(r'\s+', ' ', str(val).strip())


def _norm_header(val):
    text = _norm(val).lower()
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return ' '.join(text.split())


def _parse_batch(div_val):
    """A1(IPE) -> A1, B3(IPE) -> B3, A1 -> A1"""
    text = _norm(div_val).upper()
    m = re.match(r'^([A-Z]\d+)', text)
    return m.group(1) if m else text.split('(')[0].strip()


def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = _norm(val)
    if not text:
        return None
    # Prefer explicit date token when cell has extra text
    m = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text)
    token = m.group(1) if m else text
    token = token.replace('/', '-')
    for fmt in ('%d-%m-%Y', '%d-%m-%y', '%Y-%m-%d', '%m-%d-%Y'):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def _parse_subject_name(label):
    """FSD-2 (IPE) -> FSD-2"""
    text = _norm(label)
    text = re.sub(r'\s*\(IPE\).*$', '', text, flags=re.I)
    text = re.sub(r'\s*\(GP\).*$', '', text, flags=re.I)
    return text.strip()


def _parse_exam_type(label):
    low = _norm(label).lower()
    return 'GP' if '(gp)' in low else 'IPE'


def _match_subject(department, label):
    name = _parse_subject_name(label)
    if not name:
        return None
    qs = Subject.objects.filter(
        Q(semester=department.semester, department__isnull=True) | Q(department=department)
    )
    sub = qs.filter(name__iexact=name).first()
    if sub:
        return sub
    sub = qs.filter(code__iexact=name).first()
    if sub:
        return sub
    return qs.filter(name__icontains=name).first()


def _unique_mentor_code(department, base):
    code = re.sub(r'[^A-Za-z0-9]', '', base).upper()[:12] or 'EXT'
    if not Faculty.objects.filter(department=department, mentor_code__iexact=code).exists():
        return code
    n = 2
    while Faculty.objects.filter(department=department, mentor_code__iexact=f'{code}{n}').exists():
        n += 1
    return f'{code}{n}'


def resolve_faculty(department, identifier, *, as_external=False, create_external=True, email=''):
    """
    Match faculty by mentor code or name.
    External examiners reuse existing faculty login when matched.
    Optional email is saved/updated on the faculty record.
    """
    ident = _norm(identifier)
    email_addr = _norm(email).lower()
    if not ident:
        return None, 'empty'

    qs = Faculty.objects.filter(department=department)
    fac = qs.filter(mentor_code__iexact=ident).first()
    match_type = 'matched_code' if fac else None

    if not fac:
        fac = qs.filter(name__iexact=ident).first()
        match_type = 'matched_name' if fac else None

    if not fac and len(ident) <= 6 and ident.isupper():
        fac = qs.filter(mentor_code__iexact=ident).first()
        match_type = 'matched_code' if fac else None

    if not fac:
        parts = ident.split()
        if len(parts) >= 2:
            fac = qs.filter(
                Q(name__icontains=parts[0]) & Q(name__icontains=parts[-1])
            ).first()
            match_type = 'matched_name_partial' if fac else None

    if fac:
        if email_addr and '@' in email_addr and (fac.email or '').lower() != email_addr:
            fac.email = email_addr
            fac.save(update_fields=['email'])
        return fac, match_type or 'matched'

    if not as_external:
        return None, 'internal_not_found'

    if not create_external:
        return None, 'external_not_found'

    code = _unique_mentor_code(department, ident.replace(' ', '')[:10])
    fac = Faculty.objects.create(
        department=department,
        name=ident.title() if ident.islower() else ident,
        mentor_code=code,
        email=email_addr if '@' in email_addr else '',
        is_external=True,
    )
    return fac, 'created_external'


FLAT_HEADER_ALIASES = {
    'subject': {'subject', 'subject name', 'sub'},
    'internal': {'internal faculty', 'internal', 'internal name', 'internal examiner'},
    'external': {
        'external faculty name', 'external faculty', 'external', 'external name',
        'external examiner',
    },
    'batch': {'batch', 'div', 'division'},
    'duty_date': {'duty date', 'date', 'exam date'},
    'time_slot': {'time slot', 'time', 'slot'},
    'email': {'email id', 'email', 'e mail', 'mail', 'external email'},
    'room_no': {'room', 'room no', 'room number'},
}


def _map_flat_headers(header_row):
    mapping = {}
    for idx, raw in enumerate(header_row):
        key = _norm_header(raw)
        if not key:
            continue
        for field, aliases in FLAT_HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _find_flat_header_row(ws):
    for r in range(1, min(12, (ws.max_row or 1) + 1)):
        values = [ws.cell(r, c).value for c in range(1, (ws.max_column or 1) + 1)]
        mapping = _map_flat_headers(values)
        if 'subject' in mapping and 'batch' in mapping and (
            'external' in mapping or 'internal' in mapping
        ):
            return r, mapping
    return None, {}


def _parse_flat_schedule(ws):
    """Parse template: subject, internal faculty, External Faculty Name, Batch, Duty Date, Time Slot, Email id."""
    header_row, mapping = _find_flat_header_row(ws)
    if not header_row:
        return [], []

    rows = []
    warnings = []
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        def cell(field):
            idx = mapping.get(field)
            if idx is None:
                return None
            return ws.cell(r, idx + 1).value

        subject_label = _norm(cell('subject'))
        batch_raw = cell('batch')
        if not subject_label and not _norm(batch_raw):
            continue
        if not subject_label:
            warnings.append(f'Row {r}: missing subject')
            continue
        batch = _parse_batch(batch_raw)
        if not batch:
            warnings.append(f'Row {r}: missing batch')
            continue
        duty_date = _parse_date(cell('duty_date'))
        if not duty_date:
            warnings.append(f'Row {r}: invalid/missing duty date ({_norm(cell("duty_date"))})')
            continue

        rows.append({
            'subject_label': subject_label,
            'exam_type': _parse_exam_type(subject_label),
            'duty_date': duty_date,
            'batch': batch,
            'time_slot': _norm(cell('time_slot')),
            'room_no': _norm(cell('room_no')),
            'internal': _norm(cell('internal')),
            'external': _norm(cell('external')),
            'email': _norm(cell('email')).lower(),
        })
    return rows, warnings


def _parse_schedule_blocks(ws):
    """Detect subject/date blocks from header rows (legacy multi-column layout)."""
    blocks = []
    max_col = ws.max_column or 10
    col = 1
    while col <= max_col:
        subject_cell = ws.cell(5, col).value
        if not subject_cell or not _norm(subject_cell):
            col += 1
            continue
        date_cell = ws.cell(6, col).value
        header = _norm(ws.cell(7, col).value).lower()
        if 'div' not in header:
            col += 1
            continue
        end_col = col + 4
        if end_col > max_col:
            end_col = max_col
        blocks.append({
            'subject_col': col,
            'time_col': col + 1,
            'internal_col': col + 2,
            'external_col': col + 3,
            'room_col': col + 4,
            'subject_label': _norm(subject_cell),
            'duty_date': _parse_date(date_cell),
            'exam_type': _parse_exam_type(subject_cell),
            'data_start_row': 8,
        })
        col = end_col + 1
    return blocks


def _parse_block_schedule(ws):
    blocks = _parse_schedule_blocks(ws)
    rows = []
    warnings = []
    for block in blocks:
        if not block['duty_date']:
            warnings.append(f"Could not parse date for {block['subject_label']}")
            continue
        for r in range(block['data_start_row'], ws.max_row + 1):
            div = ws.cell(r, block['subject_col']).value
            if not div or not _norm(div):
                continue
            batch = _parse_batch(div)
            if not batch:
                continue
            rows.append({
                'subject_label': block['subject_label'],
                'exam_type': block['exam_type'],
                'duty_date': block['duty_date'],
                'batch': batch,
                'time_slot': _norm(ws.cell(r, block['time_col']).value),
                'room_no': _norm(ws.cell(r, block['room_col']).value),
                'internal': _norm(ws.cell(r, block['internal_col']).value),
                'external': _norm(ws.cell(r, block['external_col']).value),
                'email': '',
            })
    return rows, warnings


def parse_duty_schedule(file_path):
    """Parse schedule Excel into duty row dicts (no DB writes)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows, warnings = _parse_flat_schedule(ws)
    if not rows:
        rows, block_warnings = _parse_block_schedule(ws)
        warnings.extend(block_warnings)
    wb.close()
    if not rows and not warnings:
        warnings.append(
            'No duty rows found. Use columns: subject, internal faculty, '
            'External Faculty Name, Batch, Duty Date, Time Slot, Email id'
        )
    return rows, warnings


def _as_uploaded_file(uploaded_file):
    if hasattr(uploaded_file, '_committed'):
        return uploaded_file
    name = getattr(uploaded_file, 'name', 'duty_schedule.xlsx')
    content = uploaded_file.read()
    return ContentFile(content, name=name)


@transaction.atomic
def import_duty_schedule(department, uploaded_file, uploaded_by, replace_existing=False):
    """Import parsed schedule into FacultyDutyAssignment records."""
    uploaded_file = _as_uploaded_file(uploaded_file)
    upload = DutyScheduleUpload.objects.create(
        department=department,
        schedule_file=uploaded_file,
        uploaded_by=uploaded_by,
    )
    file_path = upload.schedule_file.path
    rows, parse_warnings = parse_duty_schedule(file_path)
    summary = {
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'external_created': 0,
        'external_matched': 0,
        'warnings': list(parse_warnings),
    }

    if replace_existing:
        # Permanent remove previously Excel-imported IPE duties for this department
        FacultyDutyAssignment.objects.filter(
            department=department,
            exam_type=FacultyDutyAssignment.ExamType.IPE,
            schedule_upload__isnull=False,
        ).delete()

    for row in rows:
        subject = _match_subject(department, row['subject_label'])
        if not subject:
            summary['warnings'].append(
                f"Subject not found: {row['subject_label']} (batch {row['batch']})"
            )
            summary['skipped'] += 1
            continue

        for role_key, identifier, as_external in (
            ('INTERNAL', row['internal'], False),
            ('EXTERNAL', row['external'], True),
        ):
            if not identifier:
                continue
            faculty, match_type = resolve_faculty(
                department,
                identifier,
                as_external=as_external,
                create_external=as_external,
                email=row.get('email', '') if as_external else '',
            )
            if not faculty:
                summary['warnings'].append(
                    f"{role_key} faculty not found: {identifier} "
                    f"({row['subject_label']} {row['batch']})"
                )
                summary['skipped'] += 1
                continue

            if match_type == 'created_external':
                summary['external_created'] += 1
            elif as_external and match_type and match_type.startswith('matched'):
                summary['external_matched'] += 1

            defaults = {
                'time_slot': row['time_slot'],
                'room_no': row['room_no'],
                'schedule_upload': upload,
                'assigned_by': uploaded_by,
                'is_active': True,
            }
            duty, created = FacultyDutyAssignment.objects.update_or_create(
                faculty=faculty,
                department=department,
                subject=subject,
                exam_type=row['exam_type'],
                batch=row['batch'],
                duty_date=row['duty_date'],
                duty_role=role_key,
                defaults=defaults,
            )
            if created:
                summary['created'] += 1
            else:
                summary['updated'] += 1

    upload.summary = summary
    upload.save(update_fields=['summary'])
    return upload, summary
