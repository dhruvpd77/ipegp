import io
import openpyxl
import pandas as pd
from django.http import HttpResponse


def read_excel_headers(file):
    """Read column headers from uploaded Excel/CSV."""
    if file.name.endswith('.csv'):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
    return list(df.columns), df


def import_students_from_excel(file, department):
    """Import students from standard export format."""
    from .models import Student, Faculty

    columns, df = read_excel_headers(file)
    col_map = {
        'roll no': 'roll_no', 'name': 'name', 'enrollment no': 'enrollment_no',
        'branch': 'branch', 'batch': 'batch', 'mentor': 'mentor',
        'student phone': 'student_phone', 'parents contact': 'parent_contact',
    }
    normalized = {c: col_map.get(c.lower().strip(), c.lower().strip()) for c in columns}

    created, updated, faculty_codes = 0, 0, set()
    for _, row in df.iterrows():
        data = {}
        for orig, norm in normalized.items():
            val = row.get(orig, '')
            if pd.isna(val):
                val = ''
            if norm == 'enrollment_no':
                if isinstance(val, float):
                    data[norm] = str(int(val))
                else:
                    data[norm] = str(val).strip()
            else:
                data[norm] = str(val).strip()

        if not data.get('enrollment_no'):
            continue

        mentor = data.get('mentor', '')
        if mentor:
            faculty_codes.add(mentor)

        obj, is_new = Student.objects.update_or_create(
            enrollment_no=data['enrollment_no'],
            defaults={
                'roll_no': data.get('roll_no', ''),
                'name': data.get('name', ''),
                'branch': data.get('branch', ''),
                'batch': data.get('batch', ''),
                'mentor': mentor,
                'student_phone': data.get('student_phone', ''),
                'parent_contact': data.get('parent_contact', ''),
                'department': department,
            }
        )
        if is_new:
            created += 1
        else:
            updated += 1

    for code in faculty_codes:
        if code and not Faculty.objects.filter(department=department, mentor_code=code).exists():
            Faculty.objects.create(name=code, mentor_code=code, department=department)

    return created, updated, len(faculty_codes)


def import_faculty_from_excel(file, department):
    """Import faculty list from Excel."""
    from .models import Faculty

    columns, df = read_excel_headers(file)
    created, updated = 0, 0
    for _, row in df.iterrows():
        name = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
        if not name:
            continue
        mentor_code = ''
        email = ''
        phone = ''
        for i, col in enumerate(columns):
            cl = col.lower()
            val = row[col]
            if pd.isna(val):
                continue
            val = str(val).strip()
            if 'mentor' in cl or 'code' in cl:
                mentor_code = val
            elif 'email' in cl:
                email = val
            elif 'phone' in cl:
                phone = val
            elif i == 0:
                name = val

        obj, is_new = Faculty.objects.update_or_create(
            department=department,
            name=name,
            defaults={'mentor_code': mentor_code, 'email': email, 'phone': phone}
        )
        if is_new:
            created += 1
        else:
            updated += 1
    return created, updated


def create_form_fields_from_excel(file, template, semester=None, department=None):
    """Create dynamic form fields from uploaded GP project details Excel."""
    from .gp_utils import import_cases_from_excel, import_fields_from_excel

    count, _ = import_fields_from_excel(file, template, replace=True)
    if semester:
        import_cases_from_excel(file, semester, department)
    return count


def export_credentials_excel(credentials_list, filename='credentials.xlsx'):
    """Export username/password list to Excel."""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Credentials'
    ws.append(['Username', 'Password', 'Name', 'Role'])
    for cred in credentials_list:
        ws.append(cred)
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_students_excel(students, filename='students.xlsx'):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'
    ws.append(['Roll No', 'Name', 'Enrollment No', 'Branch', 'Batch', 'Mentor', 'Student Phone', 'Parents Contact'])
    for s in students:
        ws.append([s.roll_no, s.name, s.enrollment_no, s.branch, s.batch, s.mentor, s.student_phone, s.parent_contact])
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_marksheet(student, marks):
    """Generate marksheet Excel for a student."""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Marksheet'
    ws.append(['IPE/GP Marksheet'])
    ws.append([])
    ws.append(['Name', student.name])
    ws.append(['Enrollment No', student.enrollment_no])
    ws.append(['Branch', student.branch])
    ws.append(['Batch', student.batch])
    ws.append([])
    ws.append(['Exam', 'Subject', 'Marks Obtained', 'Max Marks', 'Remarks'])
    for m in marks:
        max_m = m.exam_session.subject.max_marks_ipe if m.exam_session.exam_type == 'IPE' else m.exam_session.subject.max_marks_gp
        ws.append([m.exam_session.exam_type, m.exam_session.subject.name, float(m.marks_obtained), max_m, m.remarks])
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="marksheet_{student.enrollment_no}.xlsx"'
    return response


def roll_no_sort_key(roll_no):
    """Numeric ascending sort key for roll numbers stored as strings."""
    s = str(roll_no or '').strip()
    if s.isdigit():
        return (0, int(s))
    try:
        return (0, int(float(s)))
    except ValueError:
        return (1, s.lower())


def sort_students_by_roll(students):
    """Return students sorted by roll number (1, 2, 3 … not 1, 10, 11)."""
    items = list(students) if not isinstance(students, list) else students
    return sorted(items, key=lambda s: roll_no_sort_key(s.roll_no))
