"""IPE/GP marksheet template parsing and batch-wise Excel generation."""
import io
import math
import re
from pathlib import Path

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ExamSession, FacultyDutyAssignment, MarkEntry, MarksheetTemplate, Student, User
from .utils import roll_no_sort_key, sort_students_by_roll


def get_marks_by_batch(department, subject, exam_type, batch):
    """Load saved mark entries for a batch/subject from duty assignments."""
    entries = MarkEntry.objects.filter(
        student__department=department,
        student__batch=batch,
        duty_assignment__subject=subject,
        duty_assignment__exam_type=exam_type,
        duty_assignment__department=department,
        duty_assignment__is_active=True,
    ).select_related('student')
    result = {m.student_id: m for m in entries}
    if result:
        return result
    session = ExamSession.objects.filter(
        department=department, subject=subject, exam_type=exam_type, is_active=True,
    ).order_by('-created_at').first()
    if session:
        return {
            m.student_id: m
            for m in MarkEntry.objects.filter(
                exam_session=session, student__department=department, student__batch=batch,
            ).select_related('student')
        }
    return {}


def _cell_value_from_mark(col_def, mark_entry):
    """Resolve Excel cell value from stored mark_data."""
    if not mark_entry:
        return ''
    data = mark_entry.mark_data or {}
    ctype = col_def['type']
    fname = col_def.get('field_name', col_def['key'])
    if ctype == 'remarks':
        return mark_entry.remarks or data.get('remarks', '')
    if ctype == 'total':
        t = data.get('total')
        if t not in (None, '', 0, '0'):
            return t
        if mark_entry.marks_obtained and float(mark_entry.marks_obtained) != 0:
            return float(mark_entry.marks_obtained)
        return ''
    if ctype in ('mark', 'final_chit', 'deduction_aim', 'deduction_late'):
        v = data.get(fname, '')
        return v if v not in (None, '', 0, '0') else ''
    return ''

FONT_NAME = 'Times New Roman'
FONT_SIZE = 11
FONT_BOLD = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
FONT_NORMAL = Font(name=FONT_NAME, size=FONT_SIZE, bold=False)
FONT_NOTES = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='FF0000')

THIN = Side(style='thin', color='000000')
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_PEACH = PatternFill('solid', fgColor='FBD4B4')
FILL_GREY = PatternFill('solid', fgColor='D8D8D8')
FILL_HEAD = PatternFill('solid', fgColor='FDE9D9')
FILL_PURPLE = PatternFill('solid', fgColor='E5DFEC')
FILL_RED = PatternFill('solid', fgColor='FF8585')
FILL_YELLOW = PatternFill('solid', fgColor='FFFFB3')
FILL_PROG = PatternFill('solid', fgColor='F2DBDB')
FILL_TOTAL = PatternFill('solid', fgColor='EAF1DD')

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

COLUMN_FILLS = {
    'sr_no': FILL_HEAD, 'batch': FILL_HEAD, 'div': FILL_HEAD, 'group_id': FILL_HEAD,
    'enrollment_no': FILL_HEAD,
    'name': FILL_HEAD, 'roll_no': FILL_HEAD, 'final_chit': FILL_PURPLE,
    'deduction_aim': FILL_RED, 'deduction_late': FILL_YELLOW,
    'mark': FILL_PROG, 'total': FILL_TOTAL, 'remarks': FILL_HEAD,
    'gender_diversity': FILL_HEAD, 'religion_diversity': FILL_HEAD,
    'case': FILL_HEAD, 'final_marks': FILL_TOTAL,
}

GP_MARKSHEET_TITLE = 'Group Project Marksheet'


def _norm(val):
    if val is None:
        return ''
    return re.sub(r'\s+', ' ', str(val).replace('\n', ' ')).strip()


def _classify_column(label8, label9):
    low = f'{_norm(label8)} {_norm(label9)}'.lower()
    if 'sr' in low and 'no' in low:
        return 'sr_no'
    if 'group id' in low:
        return 'group_id'
    if low.strip() == 'div' or low.startswith('div '):
        return 'div'
    if low.strip() == 'batch' or low.startswith('batch '):
        return 'batch'
    if 'enrollment' in low:
        return 'enrollment_no'
    if 'name of student' in low or low.strip() == 'name':
        return 'name'
    if 'roll' in low:
        return 'roll_no'
    if 'gender' in low and 'diversity' in low:
        return 'gender_diversity'
    if 'religion' in low and 'diversity' in low:
        return 'religion_diversity'
    if 'final' in low and 'mark' in low:
        return 'final_marks'
    if 'case' in low and ('1' in low or '2' in low or '3' in low or '/' in low):
        return 'case'
    if 'final' in low and 'chit' in low:
        return 'final_chit'
    if 'change' in low and 'aim' in low:
        return 'deduction_aim'
    if 'late' in low:
        return 'deduction_late'
    if 'remark' in low:
        return 'remarks'
    if 'program' in low or 'logic' in low or 'output' in low or 'ppt' in low or 'presentation' in low:
        return 'mark'
    if 'total' in low and 'mark' in low:
        return 'total'
    if ('ipe:' in low or 'gp:' in low) or (
        'total' in low and ('ipe' in low or 'gp' in low) and 'program' not in low
    ):
        return 'total'
    return 'mark'


def _extract_max_marks(label):
    m = re.search(r'\((\d+)\)', str(label or ''))
    return int(m.group(1)) if m else None


def _find_header_row(ws):
    for row in range(1, 25):
        texts = [_norm(ws.cell(row, c).value).lower() for c in range(1, ws.max_column + 1)]
        if any('enrollment' in t for t in texts) and any('sr' in t for t in texts):
            return row
    return 8


def _is_gp_marksheet_format(ws, header_row):
    for col in range(1, ws.max_column + 1):
        val = _norm(ws.cell(header_row, col).value).lower()
        if 'group id' in val:
            return True
    return False


def _find_gp_meta_start(ws, header_row):
    for col in range(7, ws.max_column + 1):
        val = _norm(ws.cell(header_row, col).value).lower()
        if 'gender' in val and 'diversity' in val:
            return col
    return 12


def _parse_gp_marksheet_schema(ws, header_row):
    """Parse GP group-project marksheet with 3-row headers (attendance-style groups)."""
    mark_row = header_row + 1
    marks_sub_row = header_row + 2
    data_start = marks_sub_row + 1
    meta_start = _find_gp_meta_start(ws, header_row)
    last_col = max(ws.max_column, meta_start + 3)

    std_keys = ['sr_no', 'div', 'group_id', 'roll_no', 'enrollment_no', 'name']
    columns = []
    for col in range(1, 7):
        label8 = ws.cell(header_row, col).value
        columns.append({
            'col': col,
            'key': std_keys[col - 1],
            'field_name': std_keys[col - 1],
            'type': std_keys[col - 1],
            'label8': str(label8) if label8 else '',
            'label9': '',
            'marks_sub_label': '',
            'max_marks': None,
        })

    for col in range(7, meta_start):
        label_mr = ws.cell(mark_row, col).value
        label_sr = ws.cell(marks_sub_row, col).value
        key = _classify_column('', label_mr or label_sr)
        col_type = 'total' if key == 'total' else 'mark'
        columns.append({
            'col': col,
            'key': key,
            'field_name': f'mark_{col}',
            'type': col_type,
            'label8': '',
            'label9': str(label_mr) if label_mr else '',
            'marks_sub_label': str(label_sr) if label_sr else '',
            'max_marks': _extract_max_marks(label_sr) or _extract_max_marks(label_mr),
        })

    for col in range(meta_start, last_col + 1):
        label8 = ws.cell(header_row, col).value
        label_sr = ws.cell(marks_sub_row, col).value
        key = _classify_column(label8, label_sr)
        col_type = key if key in (
            'gender_diversity', 'religion_diversity', 'case', 'final_marks',
        ) else 'mark'
        columns.append({
            'col': col,
            'key': key,
            'field_name': key,
            'type': col_type,
            'label8': str(label8) if label8 else '',
            'label9': '',
            'marks_sub_label': str(label_sr) if label_sr else '',
            'max_marks': _extract_max_marks(label_sr) or _extract_max_marks(label8),
        })

    merges = []
    for m in ws.merged_cells.ranges:
        if m.min_row >= header_row and m.max_row <= marks_sub_row:
            merges.append({
                'range': str(m),
                'min_row': m.min_row, 'max_row': m.max_row,
                'min_col': m.min_col, 'max_col': m.max_col,
            })

    mark_cols = [c['col'] for c in columns if c['type'] == 'mark']
    total_col = next((c['col'] for c in columns if c['type'] == 'total'), None)
    sum_cols = mark_cols[:] if mark_cols else []
    if total_col and total_col not in sum_cols:
        pass
    elif total_col:
        sum_cols = [c for c in mark_cols if c < total_col]

    widths = {}
    for c in columns:
        letter = get_column_letter(c['col'])
        w = ws.column_dimensions[letter].width
        if w:
            widths[str(c['col'])] = round(w, 2)

    return {
        'gp_format': True,
        'header_row': header_row,
        'mark_row': mark_row,
        'marks_sub_row': marks_sub_row,
        'subheader_row': mark_row,
        'data_start_row': data_start,
        'columns': columns,
        'merges': merges,
        'mark_cols': mark_cols,
        'sum_cols': sum_cols or mark_cols,
        'total_col': total_col,
        'gender_col': next((c['col'] for c in columns if c['key'] == 'gender_diversity'), None),
        'religion_col': next((c['col'] for c in columns if c['key'] == 'religion_diversity'), None),
        'case_col': next((c['col'] for c in columns if c['key'] == 'case'), None),
        'final_col': next((c['col'] for c in columns if c['key'] == 'final_marks'), None),
        'widths': widths,
        'last_col': max((c['col'] for c in columns), default=15),
    }


def _parse_standard_marksheet_schema(ws, header_row):
    """Parse standard IPE-style marksheet (per-student rows)."""
    subheader_row = header_row + 1
    has_sub = any(ws.cell(subheader_row, c).value for c in range(1, ws.max_column + 1))
    if not has_sub:
        subheader_row = None
    data_start = (subheader_row or header_row) + 1

    parent_l8 = ''
    columns = []
    for col in range(1, ws.max_column + 1):
        l8 = ws.cell(header_row, col).value
        l9 = ws.cell(subheader_row, col).value if subheader_row else None
        if l8:
            parent_l8 = str(l8)
        if not l8 and not l9:
            continue
        eff8 = str(l8) if l8 else parent_l8
        key = _classify_column(eff8, l9)
        col_type = 'mark' if key == 'mark' else key
        field_name = key if key != 'mark' else f'mark_{col}'
        columns.append({
            'col': col,
            'key': key,
            'field_name': field_name,
            'type': col_type,
            'label8': str(l8) if l8 else '',
            'label9': str(l9) if l9 else '',
            'parent8': eff8 if not l8 else '',
            'max_marks': _extract_max_marks(l9) or _extract_max_marks(l8) or _extract_max_marks(eff8),
        })

    merges = []
    for m in ws.merged_cells.ranges:
        if m.min_row >= header_row and m.max_row <= (subheader_row or header_row):
            merges.append({
                'range': str(m),
                'min_row': m.min_row, 'max_row': m.max_row,
                'min_col': m.min_col, 'max_col': m.max_col,
            })

    mark_cols = [c['col'] for c in columns if c['type'] == 'mark']
    total_col = next((c['col'] for c in columns if c['type'] == 'total'), None)
    sum_cols = [
        c['col'] for c in columns
        if c['type'] in ('mark', 'deduction_aim', 'deduction_late')
    ]
    if not sum_cols and total_col:
        sum_cols = [c['col'] for c in columns if c['col'] < total_col and c['type'] not in (
            'sr_no', 'batch', 'enrollment_no', 'name', 'roll_no', 'remarks', 'total',
        )]
    widths = {}
    for c in columns:
        letter = get_column_letter(c['col'])
        w = ws.column_dimensions[letter].width
        if w:
            widths[str(c['col'])] = round(w, 2)

    return {
        'header_row': header_row,
        'subheader_row': subheader_row,
        'data_start_row': data_start,
        'columns': columns,
        'merges': merges,
        'mark_cols': mark_cols,
        'sum_cols': sum_cols,
        'total_col': total_col,
        'widths': widths,
        'last_col': max((c['col'] for c in columns), default=14),
    }


def parse_marksheet_template(file):
    """Parse uploaded marksheet Excel into a JSON schema (columns + merges)."""
    wb = openpyxl.load_workbook(file, data_only=False)
    ws = wb.active
    header_row = _find_header_row(ws)
    if _is_gp_marksheet_format(ws, header_row):
        schema = _parse_gp_marksheet_schema(ws, header_row)
    else:
        schema = _parse_standard_marksheet_schema(ws, header_row)
    wb.close()
    return schema


def resolve_marksheet_template(exam_type, subject, department):
    semester = department.semester if department else subject.semester
    return MarksheetTemplate.objects.filter(
        exam_type=exam_type, subject=subject,
    ).filter(
        Q(department=department) | Q(department__isnull=True, semester=semester)
    ).order_by('-department_id', '-created_at').first()


def _fill_for_column(col_def):
    return COLUMN_FILLS.get(col_def['type'], COLUMN_FILLS.get(col_def['key'], FILL_HEAD))


def _style(cell, font=None, fill=None, alignment=None):
    cell.font = font or FONT_NORMAL
    if fill:
        cell.fill = fill
    cell.alignment = alignment or ALIGN_C
    cell.border = BORDER_THIN


def _merge_header(ws, r1, c1, r2, c2, value, fill=None, alignment=None):
    aln = alignment or ALIGN_C
    f = fill or FILL_HEAD
    if r1 == r2 and c1 == c2:
        cell = ws.cell(r1, c1, value)
        _style(cell, font=FONT_BOLD, fill=f, alignment=aln)
        return
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1, value)
    _style(cell, font=FONT_BOLD, fill=f, alignment=aln)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            _style(ws.cell(r, c), font=FONT_BOLD, fill=f, alignment=aln)


def _build_headers_from_schema(ws, schema, last_col):
    """Rebuild two-row table headers from parsed schema using stored merges."""
    hr = schema['header_row']
    shr = schema.get('subheader_row')

    for col_def in schema['columns']:
        c = col_def['col']
        fill = _fill_for_column(col_def)
        if col_def['label8']:
            _style(ws.cell(hr, c, col_def['label8']), font=FONT_BOLD, fill=fill)
        if shr and col_def['label9']:
            sub_fill = FILL_PROG if col_def['type'] == 'mark' else fill
            _style(ws.cell(shr, c, col_def['label9']), font=FONT_BOLD, fill=sub_fill)

    for m in schema.get('merges', []):
        try:
            ws.merge_cells(m['range'])
        except ValueError:
            pass

    for m in schema.get('merges', []):
        cell = ws.cell(m['min_row'], m['min_col'])
        col_def = next((c for c in schema['columns'] if c['col'] == m['min_col']), None)
        if col_def:
            _style(cell, font=FONT_BOLD, fill=_fill_for_column(col_def))


# IPE portrait layout — narrow deduction cols; program + total cols fixed wider.
IPE_FIXED_WIDTHS = {
    'sr_no': 6,
    'batch': 6,
    'roll_no': 7,
    'final_chit': 8,
    'deduction_aim': 9,
    'deduction_late': 10,
    'mark': 12,
    'total': 12,
}

IPE_DYNAMIC_LIMITS = {
    'enrollment_no': (14, 17),
    'name': (26, 34),
    'remarks': (10, 14),
}


def _content_max_len(ws, col, rows):
    max_len = 0
    for row in rows:
        if row is None:
            continue
        val = ws.cell(row, col).value
        if val is not None:
            max_len = max(max_len, len(str(val)))
    return max_len


def _apply_marksheet_column_widths(ws, schema, data_start, data_end):
    """Portrait-friendly widths: narrow deductions, proper program/total cols, dynamic rest."""
    hr = schema['header_row']
    shr = schema.get('subheader_row')
    header_rows = [r for r in (hr, shr) if r]

    for col_def in schema['columns']:
        col = col_def['col']
        letter = get_column_letter(col)
        ctype = col_def['type']
        key = col_def['key']

        if ctype in IPE_FIXED_WIDTHS:
            width = IPE_FIXED_WIDTHS[ctype]
        elif key in IPE_DYNAMIC_LIMITS:
            min_w, max_w = IPE_DYNAMIC_LIMITS[key]
            scan_rows = header_rows + list(range(data_start, data_end + 1))
            max_len = _content_max_len(ws, col, scan_rows)
            width = min(max(max_len + 1, min_w), max_w)
        else:
            scan_rows = header_rows + list(range(data_start, data_end + 1))
            max_len = _content_max_len(ws, col, scan_rows)
            width = min(max(max_len + 1, 8), 14)

        ws.column_dimensions[letter].width = width

        if ctype in ('deduction_aim', 'deduction_late'):
            for row in header_rows:
                cell = ws.cell(row, col)
                if cell.value:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center', wrap_text=True,
                    )


def _apply_marksheet_row_heights(ws, schema, data_start, data_end):
    """Row heights for headers (wrapped deductions) and student data."""
    hr = schema['header_row']
    shr = schema.get('subheader_row')
    name_col = next((c['col'] for c in schema['columns'] if c['key'] == 'name'), 4)

    ws.row_dimensions[7].height = 40
    ws.row_dimensions[hr].height = 36
    if shr:
        ws.row_dimensions[shr].height = 30

    name_width = ws.column_dimensions[get_column_letter(name_col)].width or 28
    num_rows = max(data_end - data_start + 1, 1)
    base_h = 15 if num_rows > 25 else 16
    for row in range(data_start, data_end + 1):
        name_val = ws.cell(row, name_col).value
        if name_val and len(str(name_val)) > int(name_width * 0.85):
            ws.row_dimensions[row].height = base_h + 3
        else:
            ws.row_dimensions[row].height = base_h


def _total_sheet_height(ws, last_row):
    return sum(ws.row_dimensions[r].height or 15 for r in range(1, last_row + 1))


def _calc_portrait_print_scale(ws, last_col, last_row):
    """Scale % so all columns fit width on A4 portrait (official template ~22%)."""
    col_pt_factor = 7.0
    total_w = sum(
        (ws.column_dimensions[get_column_letter(c)].width or 8) * col_pt_factor
        for c in range(1, last_col + 1)
    )
    if total_w <= 0:
        return 22
    page_w = 559
    return max(12, min(int(page_w / total_w * 100), 100))


def _expand_rows_to_fill_page(ws, last_row, scale_pct, page_h=806):
    """Grow row heights so scaled content fills the full portrait page height."""
    if scale_pct <= 0:
        return
    total_h = _total_sheet_height(ws, last_row)
    target_h = page_h / (scale_pct / 100)
    if total_h >= target_h * 0.95:
        return
    factor = target_h / total_h
    for r in range(1, last_row + 1):
        h = ws.row_dimensions[r].height or 15
        ws.row_dimensions[r].height = min(h * factor, 45)


def _apply_marksheet_print_setup(ws, last_col, last_row):
    """Portrait A4 — all rows and columns on one page, filling the full sheet."""
    from openpyxl.worksheet.page import PageMargins

    last_letter = get_column_letter(last_col)
    scale = _calc_portrait_print_scale(ws, last_col, last_row)
    _expand_rows_to_fill_page(ws, last_row, scale)

    ws.print_area = f'A1:{last_letter}{last_row}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.scale = scale
    ws.page_setup.fitToWidth = None
    ws.page_setup.fitToHeight = None
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.3, bottom=0.3, header=0.1, footer=0.1,
    )
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False


def _student_cell_value(key, student, sr_no, batch):
    if key == 'sr_no':
        return sr_no
    if key == 'batch':
        return batch
    if key == 'enrollment_no':
        return str(student.enrollment_no or '')
    if key == 'name':
        return (student.name or '').strip().upper()
    if key == 'roll_no':
        return str(student.roll_no or '')
    return ''


def _build_batch_sheet(
    ws, schema, batch, students, subject, exam_type,
    semester_label, department_label, marks_by_student,
    per_student_batch=False,
):
    last_col = schema['last_col']
    last_letter = get_column_letter(last_col)
    hr = schema['header_row']
    data_start = schema['data_start_row']

    title = f'{exam_type} Marksheet Of Batch {batch}' if not per_student_batch else f'{exam_type} Marksheet — All Batches (Combined)'
    _merge_header(ws, 1, 1, 1, last_col, 'L. J. University', fill=FILL_PEACH)
    _merge_header(ws, 2, 1, 2, last_col, 'L. J. Institute of Engineering & Technology, Ahmedabad', fill=FILL_PEACH)
    _merge_header(ws, 3, 1, 3, last_col, f'Semester - {semester_label}   {department_label} Department', fill=FILL_PEACH)
    _merge_header(ws, 4, 1, 4, last_col, title, fill=FILL_GREY)

    date_start_col = min(9, last_col)
    _merge_header(ws, 5, 1, 5, date_start_col - 1, f'SUBJECT NAME: {subject.name}', alignment=ALIGN_L)
    _merge_header(ws, 5, date_start_col, 5, last_col, 'DATE :', alignment=ALIGN_L)
    _merge_header(ws, 6, 1, 6, date_start_col - 1, f'SUBJECT CODE:{subject.code or ""}', alignment=ALIGN_L)
    _merge_header(ws, 6, date_start_col, 6, last_col, 'START TIME:                        END TIME:', alignment=ALIGN_L)

    notes = (
        'N.B : \n'
        '1) For absent students, enter "AB" in each marks column.\n'
        '2) Do not merge any entries.'
    )
    _merge_header(ws, 7, 1, 7, last_col, notes)
    ws.cell(7, 1).font = FONT_NOTES
    ws.cell(7, 1).alignment = ALIGN_L

    _build_headers_from_schema(ws, schema, last_col)

    sum_cols = schema.get('sum_cols') or schema.get('mark_cols') or [
        c['col'] for c in schema['columns'] if c['type'] == 'mark'
    ]
    total_col = schema.get('total_col')
    remarks_col = next((c['col'] for c in schema['columns'] if c['type'] == 'remarks'), None)

    row = data_start
    for idx, student in enumerate(students, start=1):
        mark_entry = marks_by_student.get(student.pk)
        for col_def in schema['columns']:
            col = col_def['col']
            ctype = col_def['type']
            val = ''
            if ctype in ('sr_no', 'batch', 'enrollment_no', 'name', 'roll_no'):
                row_batch = student.batch if per_student_batch else batch
                val = _student_cell_value(col_def['key'], student, idx, row_batch)
            elif ctype in ('mark', 'final_chit', 'deduction_aim', 'deduction_late', 'total', 'remarks'):
                val = _cell_value_from_mark(col_def, mark_entry)
            fill = _fill_for_column(col_def)
            cell = ws.cell(row, col, val if val != '' else None)
            name_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
            _style(
                cell, fill=fill,
                alignment=name_align if col_def['key'] == 'name' else ALIGN_C,
            )

        if total_col and sum_cols:
            saved_total = _cell_value_from_mark(
                next(c for c in schema['columns'] if c['col'] == total_col), mark_entry,
            )
            if saved_total != '':
                ws.cell(row, total_col).value = saved_total
            else:
                first_l = get_column_letter(min(sum_cols))
                last_l = get_column_letter(max(sum_cols))
                ws.cell(row, total_col).value = (
                    f'=IF(SUM({first_l}{row}:{last_l}{row})=0,"",SUM({first_l}{row}:{last_l}{row}))'
                )
            _style(ws.cell(row, total_col), fill=FILL_TOTAL)

        row += 1

    data_end = row - 1
    if data_end >= data_start:
        _apply_marksheet_column_widths(ws, schema, data_start, data_end)
        _apply_marksheet_row_heights(ws, schema, data_start, data_end)

    footer_start = row + 1
    footers = [
        'NAME OF EXTERNAL EXAMINER: ',
        'SIGNATURE OF EXTERNAL EXAMINER:',
        'NAME OF INTERNAL EXAMINER:                     ',
        'SIGNATURE OF INTERNAL EXAMINER:',
    ]
    for i, text in enumerate(footers):
        fr = footer_start + i
        _merge_header(ws, fr, 1, fr, last_col - 1 if last_col > 1 else last_col, text, fill=None)
        ws.cell(fr, 1).alignment = ALIGN_L

    footer_end = footer_start + len(footers) - 1
    for fr in range(footer_start, footer_end + 1):
        ws.row_dimensions[fr].height = 16

    _apply_marksheet_print_setup(ws, last_col, footer_end)


def _ensure_schema(template):
    schema = template.schema or {}
    needs_parse = not schema.get('columns')
    if template.exam_type == 'GP' and not schema.get('gp_format'):
        needs_parse = True
    if needs_parse:
        schema = parse_marksheet_template(template.template_file.path)
        template.schema = schema
        template.save(update_fields=['schema'])
    return schema


def _gp_group_value(col_def, group, mem_idx, sr_no, batch_name, gid,
                    gender_code, religion_code, case_code, student, mark_entry):
    """Cell value for one GP marksheet data row."""
    key = col_def['key']
    ctype = col_def['type']
    if key == 'sr_no':
        return sr_no
    if key == 'div':
        return batch_name if mem_idx == 0 else None
    if key == 'group_id':
        return gid if mem_idx == 0 else None
    if key == 'roll_no':
        return str(student.roll_no or '')
    if key == 'enrollment_no':
        return str(student.enrollment_no or '')
    if key == 'name':
        return (student.name or '').strip().upper()
    if key == 'gender_diversity':
        return gender_code if mem_idx == 0 else None
    if key == 'religion_diversity':
        return religion_code if mem_idx == 0 else None
    if key == 'case':
        return case_code if mem_idx == 0 else None
    if ctype in ('mark', 'final_chit', 'deduction_aim', 'deduction_late', 'total', 'final_marks', 'remarks'):
        return _cell_value_from_mark(col_def, mark_entry)
    return ''


def _build_gp_marksheet_headers(ws, schema, last_col):
    """Rebuild 3-row GP marksheet table headers from parsed schema."""
    hr = schema['header_row']
    mr = schema.get('mark_row', hr + 1)
    sr = schema.get('marks_sub_row', hr + 2)

    for col_def in schema['columns']:
        col = col_def['col']
        fill = _fill_for_column(col_def)
        if col_def.get('label8'):
            _style(ws.cell(hr, col, col_def['label8']), font=FONT_BOLD, fill=fill)
        if col_def.get('label9'):
            sub_fill = FILL_PROG if col_def['type'] == 'mark' else fill
            _style(ws.cell(mr, col, col_def['label9']), font=FONT_BOLD, fill=sub_fill)
        if col_def.get('marks_sub_label'):
            sub_fill = FILL_PROG if col_def['type'] in ('mark', 'total') else fill
            _style(ws.cell(sr, col, col_def['marks_sub_label']), font=FONT_BOLD, fill=sub_fill)

    for m in schema.get('merges', []):
        try:
            ws.merge_cells(m['range'])
        except ValueError:
            pass

    for m in schema.get('merges', []):
        cell = ws.cell(m['min_row'], m['min_col'])
        col_def = next((c for c in schema['columns'] if c['col'] == m['min_col']), None)
        if col_def:
            _style(cell, font=FONT_BOLD, fill=_fill_for_column(col_def))


# Landscape A4 GP marksheet — tuned so all ~15 columns fit one printable page.
GP_MARKSHEET_WIDTH_LIMITS = {
    'sr_no': (4, 5),
    'div': (4, 6),
    'group_id': (8, 11),
    'roll_no': (5, 7),
    'enrollment_no': (13, 16),
    'name': (24, 32),
    'mark': (6, 9),
    'total': (7, 9),
    'gender_diversity': (8, 10),
    'religion_diversity': (8, 10),
    'case': (8, 11),
    'final_marks': (9, 12),
}


def _autofit_gp_marksheet_columns(ws, schema, data_start, data_end):
    """Compact landscape column widths so the whole marksheet fits one page."""
    for col_def in schema['columns']:
        col = col_def['col']
        key = col_def['key']
        min_w, max_w = GP_MARKSHEET_WIDTH_LIMITS.get(key, (7, 10))
        max_len = 0
        for row in range(data_start, data_end + 1):
            val = ws.cell(row, col).value
            if val is not None and not (isinstance(val, str) and val.startswith('=')):
                max_len = max(max_len, len(str(val)))
        # Headers wrap (multi-line), so don't let long header labels widen the column.
        width = min(max(max_len + 1, min_w), max_w)
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_gp_marksheet_sheet(
    ws, schema, groups, batch_name,
    semester_label, department_label,
    subject_name, subject_code,
    marks_by_student=None,
    group_id_map=None,
):
    """Build one GP marksheet sheet per formation split (attendance-style group layout)."""
    from .attendance_sheet import (
        GP_FILL_DIV,
        GP_FILL_GROUP_ALT,
        GP_FILL_PEACH,
        GP_FILL_TBL_HEAD,
        THICK_SIDE,
        THIN_SIDE,
        _apply_gp_print_setup,
        _autofit_gp_row_heights,
        _gp_cell_str,
        _gp_display_name,
        _merge_and_style,
        _style_cell,
        ALIGN_CENTER,
        ALIGN_LEFT,
        FILL_GREY,
        FONT_GP_BOLD,
        FONT_GP_NORMAL,
        FONT_GP_NOTES,
    )
    from .gp_utils import (
        gp_case_code,
        gp_gender_diversity_code,
        gp_religion_diversity_code,
        _batch_palette,
        _read_group_id,
    )
    from .models import GPGroupMemberDetail

    marks_by_student = marks_by_student or {}
    last_col = schema['last_col']
    last_letter = get_column_letter(last_col)
    hr = schema['header_row']
    data_start = schema['data_start_row']
    name_col = next((c['col'] for c in schema['columns'] if c['key'] == 'name'), 6)
    sum_cols = schema.get('sum_cols') or schema.get('mark_cols') or []
    total_col = schema.get('total_col')
    group_merge_cols = {
        c['col'] for c in schema['columns']
        if c['key'] in ('group_id', 'gender_diversity', 'religion_diversity', 'case')
    }

    ws.sheet_view.showGridLines = True

    _merge_and_style(ws, f'A1:{last_letter}1', 'L. J. University', font=FONT_GP_BOLD, fill=GP_FILL_PEACH)
    _merge_and_style(
        ws, f'A2:{last_letter}2',
        'L. J. Institute of Engineering & Technology, Ahmedabad',
        font=FONT_GP_BOLD, fill=GP_FILL_PEACH,
    )
    _merge_and_style(
        ws, f'A3:{last_letter}3',
        f'Semester - {semester_label}   {department_label} Department',
        font=FONT_GP_BOLD, fill=GP_FILL_PEACH,
    )
    _merge_and_style(ws, f'A4:{last_letter}4', GP_MARKSHEET_TITLE, font=FONT_GP_BOLD, fill=FILL_GREY)

    date_start_col = get_column_letter(max(last_col - 5, 7))
    _merge_and_style(
        ws, f'A5:{get_column_letter(max(last_col - 6, 6))}5',
        f'SUBJECT NAME: {subject_name}',
        font=FONT_GP_BOLD, alignment=ALIGN_LEFT,
    )
    _merge_and_style(ws, f'{date_start_col}5:{last_letter}5', 'Date:', font=FONT_GP_BOLD, alignment=ALIGN_LEFT)
    _merge_and_style(
        ws, f'A6:{get_column_letter(max(last_col - 6, 6))}6',
        f'SUBJECT CODE: {subject_code}',
        font=FONT_GP_BOLD, alignment=ALIGN_LEFT,
    )
    _merge_and_style(
        ws, f'{date_start_col}6:{last_letter}6',
        'Start Time:                                   End Time:',
        font=FONT_GP_BOLD, alignment=ALIGN_LEFT,
    )

    notes = (
        'N.B : \n'
        '1) For absent students, enter "AB" in each marks column.\n'
        '2) Do not merge any entries.'
    )
    _merge_and_style(ws, f'A7:{last_letter}7', notes, font=FONT_GP_NOTES, alignment=ALIGN_LEFT)

    _build_gp_marksheet_headers(ws, schema, last_col)

    all_batches = {batch_name}
    light, dark = _batch_palette(batch_name, all_batches)
    data_row = data_start
    sr_no = 1
    merge_ranges = []
    group_end_rows = []
    group_idx = 0

    for group in groups:
        group_idx += 1
        details = list(group.member_details.select_related('student').order_by('student__roll_no'))
        if not details:
            for m in group.members.order_by('roll_no'):
                details.append(GPGroupMemberDetail(student=m, gender='', member_data={}))
        if not details:
            continue

        gid = (group_id_map or {}).get(group)
        if not gid:
            gid = _read_group_id(group) or f'{batch_name}_{group_idx}'
        gender_code = gp_gender_diversity_code(group)
        religion_code = gp_religion_diversity_code(group)
        case_code = gp_case_code(group)
        start_row = data_row
        group_fill = PatternFill('solid', fgColor=dark if group_idx % 2 == 0 else light)

        for mem_idx, detail in enumerate(details):
            student = detail.student
            mark_entry = marks_by_student.get(student.pk)
            is_last = mem_idx == len(details) - 1
            row_border = Border(
                left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE,
                bottom=THICK_SIDE if is_last else THIN_SIDE,
            )

            for col_def in schema['columns']:
                col = col_def['col']
                val = _gp_group_value(
                    col_def, group, mem_idx, sr_no, batch_name, gid,
                    gender_code, religion_code, case_code, student, mark_entry,
                )
                if val is None:
                    continue
                if col_def['key'] == 'roll_no':
                    val = _gp_cell_str(val)
                elif col_def['key'] == 'enrollment_no':
                    val = _gp_cell_str(val)
                elif col_def['key'] == 'name':
                    val = _gp_display_name(student)

                cell = ws.cell(row=data_row, column=col, value=val if val != '' else None)
                fill = group_fill
                if col_def['key'] == 'div' and val:
                    fill = GP_FILL_DIV
                elif col_def['key'] == 'group_id' and val:
                    fill = GP_FILL_GROUP_ALT
                elif col_def['type'] == 'mark':
                    fill = FILL_PROG
                elif col_def['type'] == 'total':
                    fill = FILL_TOTAL
                elif col_def['key'] == 'final_marks':
                    fill = FILL_TOTAL
                else:
                    fill = _fill_for_column(col_def)
                align = ALIGN_LEFT if col == name_col else ALIGN_CENTER
                _style_cell(
                    cell, font=FONT_GP_NORMAL, fill=fill,
                    alignment=align, border_obj=row_border,
                )

            if total_col and sum_cols:
                saved_total = _cell_value_from_mark(
                    next(c for c in schema['columns'] if c['col'] == total_col),
                    mark_entry,
                )
                if saved_total != '':
                    ws.cell(data_row, total_col).value = saved_total
                else:
                    first_l = get_column_letter(min(sum_cols))
                    last_l = get_column_letter(max(sum_cols))
                    ws.cell(data_row, total_col).value = (
                        f'=IF(SUM({first_l}{data_row}:{last_l}{data_row})=0,"",'
                        f'SUM({first_l}{data_row}:{last_l}{data_row}))'
                    )
                _style_cell(ws.cell(data_row, total_col), fill=FILL_TOTAL, border_obj=row_border)

            sr_no += 1
            data_row += 1

        end_row = data_row - 1
        if end_row >= start_row:
            for col in group_merge_cols:
                merge_ranges.append((start_row, end_row, col))
            group_end_rows.append(end_row)

    data_end = data_row - 1
    div_col = next((c['col'] for c in schema['columns'] if c['key'] == 'div'), 2)
    if data_end >= data_start:
        ws.merge_cells(start_row=data_start, start_column=div_col, end_row=data_end, end_column=div_col)
        div_cell = ws.cell(data_start, div_col, batch_name)
        _style_cell(div_cell, fill=GP_FILL_DIV, alignment=ALIGN_CENTER)

    for start_row, end_row, col in merge_ranges:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        merged = ws.cell(row=start_row, column=col)
        merged.fill = GP_FILL_GROUP_ALT if col == next(
            (c['col'] for c in schema['columns'] if c['key'] == 'group_id'), 3,
        ) else _fill_for_column(next(c for c in schema['columns'] if c['col'] == col))
        merged.alignment = ALIGN_CENTER

    for end_row in group_end_rows:
        for col in range(1, last_col + 1):
            cell = ws.cell(row=end_row, column=col)
            cell.border = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THICK_SIDE)

    footer_start = data_end + 1 if data_end >= data_start else data_start
    footers = [
        'NAME OF EXTERNAL EXAMINER: ',
        'SIGNATURE OF EXTERNAL EXAMINER:',
        'NAME OF INTERNAL EXAMINER:',
        'SIGNATURE OF INTERNAL EXAMINER:',
    ]
    footer_end = footer_start + len(footers) - 1
    for i, text in enumerate(footers):
        row = footer_start + i
        _merge_and_style(ws, f'A{row}:{last_letter}{row}', text, font=FONT_GP_NORMAL, alignment=ALIGN_LEFT)
        ws.row_dimensions[row].height = 16

    for r in range(1, 5):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 16

    if data_end >= data_start:
        _autofit_gp_marksheet_columns(ws, schema, data_start=data_start, data_end=data_end)
        _autofit_gp_row_heights(ws, name_col, header_row=hr, data_start=data_start, data_end=data_end, notes_row=7)
        # Header block (3 rows) needs enough height for wrapped multi-line labels.
        ws.row_dimensions[hr].height = 30
        ws.row_dimensions[schema.get('mark_row', hr + 1)].height = 26
        ws.row_dimensions[schema.get('marks_sub_row', hr + 2)].height = 58

    _apply_gp_print_setup(ws, last_col, footer_end)


def generate_gp_marksheet_workbook(
    template,
    department,
    semester_label,
    department_label,
    subject_selection=None,
):
    """GP marksheet — one sheet per saved formation split (group layout like attendance)."""
    from .group_formation import (
        formation_key_for_selection,
        iter_formation_sheets,
        primary_subject_for_selection,
        resolve_subject_selection,
    )

    schema = _ensure_schema(template)
    group_selection = resolve_subject_selection(
        department, subject=template.subject, subject_selection=subject_selection,
    )
    formation_selection = formation_key_for_selection(department, group_selection)
    display_subject = primary_subject_for_selection(department, group_selection) or template.subject
    subject_name = display_subject.name if display_subject else 'GP'
    subject_code = (display_subject.code or '') if display_subject else ''

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_count = 0
    used_titles = set()

    for sheet_title, batch, split_index, groups, group_to_gid in iter_formation_sheets(
        department, formation_selection, group_subject_selection=group_selection,
    ):
        marks_by_student = {}
        for group in groups:
            details = list(group.member_details.select_related('student'))
            if not details:
                for m in group.members.all():
                    details.append(type('Detail', (), {'student': m})())
            for detail in details:
                stu = detail.student
                marks_by_student.update(
                    get_marks_by_batch(department, template.subject, template.exam_type, stu.batch),
                )

        unique_title = sheet_title
        suffix = 2
        while unique_title in used_titles:
            tail = f' ({suffix})'
            unique_title = sheet_title[: 31 - len(tail)] + tail
            suffix += 1
        used_titles.add(unique_title)
        ws = wb.create_sheet(title=unique_title)
        _build_gp_marksheet_sheet(
            ws, schema, groups, batch,
            semester_label, department_label,
            subject_name, subject_code,
            marks_by_student=marks_by_student,
            group_id_map=group_to_gid,
        )
        sheet_count += 1

    if not sheet_count:
        ws = wb.create_sheet(title='No Data')
        ws.cell(
            1, 1,
            'No group formation splits found. Configure splits under Group Formations first.',
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe = (template.subject.code or template.subject.name).replace(' ', '_')
    filename = f'Marksheet_GP_{safe}_{department.name}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_marksheet_workbook(
    template,
    department,
    semester_label,
    department_label,
    combined=False,
    subject_selection=None,
):
    """Generate marksheet Excel — one sheet per batch, or formation splits for GP."""
    if template.exam_type == 'GP':
        return generate_gp_marksheet_workbook(
            template, department, semester_label, department_label,
            subject_selection=subject_selection,
        )

    schema = _ensure_schema(template)
    students_qs = Student.objects.filter(department=department).order_by('batch', 'roll_no')
    batches = list(students_qs.values_list('batch', flat=True).distinct().order_by('batch'))
    if not batches:
        batches = ['A1']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if combined:
        all_students = sort_students_by_roll(students_qs)
        marks_by_student = {}
        for batch in batches:
            marks_by_student.update(
                get_marks_by_batch(department, template.subject, template.exam_type, batch)
            )
        ws = wb.create_sheet(title='COMBINE')
        _build_batch_sheet(
            ws, schema, 'ALL', all_students, template.subject,
            template.exam_type, semester_label, department_label, marks_by_student,
            per_student_batch=True,
        )
    else:
        for batch in batches:
            batch_students = sort_students_by_roll(students_qs.filter(batch=batch))
            marks_by_student = get_marks_by_batch(
                department, template.subject, template.exam_type, batch,
            )
            ws = wb.create_sheet(title=str(batch)[:31])
            _build_batch_sheet(
                ws, schema, batch, batch_students, template.subject,
                template.exam_type, semester_label, department_label, marks_by_student,
            )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe = (template.subject.code or template.subject.name).replace(' ', '_')
    suffix = '_COMBINE' if combined else ''
    filename = f'Marksheet_{template.exam_type}_{safe}_{department.name}{suffix}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _sort_students_by_batch_and_roll(students):
    items = list(students) if not isinstance(students, list) else students
    return sorted(items, key=lambda s: (s.batch or '', roll_no_sort_key(s.roll_no)))


def _compile_field_names(schema):
    """Resolve aim/late/total field names from marksheet schema."""
    aim_f = late_f = None
    if schema and schema.get('columns'):
        for col in schema['columns']:
            fname = col.get('field_name', col['key'])
            if col['type'] == 'deduction_aim':
                aim_f = fname
            elif col['type'] == 'deduction_late':
                late_f = fname
    return aim_f or 'deduction_aim', late_f or 'deduction_late'


def _estimate_wrapped_lines(text, col_width, word_wrap=True):
    """Approximate wrapped lines (word-aware) for Excel column width units."""
    if text is None or text == '':
        return 1
    max_chars = max(int((col_width or 10) * 0.78), 5)
    total = 0
    for part in str(text).split('\n'):
        part = part.strip()
        if not part:
            total += 1
            continue
        if not word_wrap or ' ' not in part:
            total += max(1, math.ceil(len(part) / max_chars))
            continue
        lines = 1
        current = 0
        for word in part.split():
            wlen = len(word)
            if current == 0:
                current = wlen
            elif current + 1 + wlen <= max_chars:
                current += 1 + wlen
            else:
                lines += 1
                current = wlen
        total += lines
    return max(total, 1)


def _compiled_row_height(max_lines):
    """Excel row height (points) for a given number of wrapped text lines."""
    line_h = 17
    return min(max(18, max_lines * line_h + 4), 72)


def _apply_compiled_row_heights(ws, header_row, data_start, data_end, last_col):
    """Dynamic row heights for wrapped enrollment, names, and deduction headers."""
    wrap_header_cols = (3, 4, 6, 7, 8)
    for col in wrap_header_cols:
        cell = ws.cell(header_row, col)
        if cell.value:
            w = ws.column_dimensions[get_column_letter(col)].width or 14
            cell.alignment = Alignment(
                horizontal='center' if col != 4 else 'left',
                vertical='center',
                wrap_text=True,
            )
    header_lines = max(
        _estimate_wrapped_lines(
            ws.cell(header_row, c).value,
            ws.column_dimensions[get_column_letter(c)].width or 10,
        )
        for c in wrap_header_cols
        if ws.cell(header_row, c).value
    )
    ws.row_dimensions[header_row].height = _compiled_row_height(header_lines)

    for row in range(data_start, data_end + 1):
        max_lines = 1
        for col in (3, 4, 6, 7):
            val = ws.cell(row, col).value
            if val is None:
                continue
            w = ws.column_dimensions[get_column_letter(col)].width or 10
            max_lines = max(max_lines, _estimate_wrapped_lines(val, w))
        ws.row_dimensions[row].height = _compiled_row_height(max_lines)


def _apply_compiled_print_setup(ws, last_col, last_row):
    """Portrait A4 — all columns and rows fit on one printed page."""
    from openpyxl.worksheet.page import PageMargins

    for r in range(1, 7):
        if not ws.row_dimensions[r].height:
            ws.row_dimensions[r].height = 16

    last_letter = get_column_letter(last_col)
    ws.print_area = f'A1:{last_letter}{last_row}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None
    ws.page_margins = PageMargins(
        left=0.2, right=0.2, top=0.25, bottom=0.25, header=0.1, footer=0.1,
    )
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False


def _compile_values_from_mark(mark_entry, aim_f, late_f):
    if not mark_entry:
        return '', '', ''
    data = mark_entry.mark_data or {}
    def _val(key):
        v = data.get(key, '')
        return '' if v in (None, '', 0, '0') else v

    aim = _val(aim_f)
    late = _val(late_f)
    total = _val('total')
    if total == '' and mark_entry.marks_obtained and float(mark_entry.marks_obtained) != 0:
        total = float(mark_entry.marks_obtained)
    return aim, late, total


def _build_compiled_sheet(
    ws, students, marks_by_student, subject, exam_type,
    semester_label, department_label, aim_f, late_f, max_marks,
):
    last_col = 8
    header_row = 7
    data_start = 8

    exam_label = exam_type.upper()
    _merge_header(ws, 1, 1, 1, last_col, 'L. J. University', fill=FILL_PEACH)
    _merge_header(
        ws, 2, 1, 2, last_col,
        'L. J. Institute of Engineering & Technology, Ahmedabad', fill=FILL_PEACH,
    )
    _merge_header(
        ws, 3, 1, 3, last_col,
        f'Semester - {semester_label}   {department_label} Department', fill=FILL_PEACH,
    )
    _merge_header(ws, 4, 1, 4, last_col, f'COMPILE {exam_label} Marksheet', fill=FILL_GREY)
    _merge_header(ws, 5, 1, 5, last_col, f'SUBJECT NAME: {subject.name}', fill=FILL_PEACH, alignment=ALIGN_L)
    _merge_header(ws, 6, 1, 6, last_col, f'SUBJECT CODE:{subject.code or ""}', fill=FILL_PEACH, alignment=ALIGN_L)

    headers = [
        (1, 'Sr No', FILL_HEAD),
        (2, 'Batch', FILL_HEAD),
        (3, 'Enrollment Number', FILL_HEAD),
        (4, 'Name of Student', FILL_HEAD),
        (5, 'Roll No', FILL_HEAD),
        (6, 'Marks Deduction Due to Change in Aim. Enter negative', FILL_RED),
        (7, 'Deduction For Late Coming, Enter negative value (-1)', FILL_YELLOW),
        (8, f'{exam_label}: Total ( {max_marks} ) Marks', FILL_TOTAL),
    ]
    for col, label, fill in headers:
        align = ALIGN_L if col == 4 else ALIGN_C
        _style(ws.cell(header_row, col, label), font=FONT_BOLD, fill=fill, alignment=align)

    # Set column widths before row-height calculation
    widths = {1: 5, 2: 6, 3: 17, 4: 30, 5: 6.5, 6: 10.5, 7: 10.5, 8: 9.5}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    row = data_start
    for idx, student in enumerate(students, start=1):
        mark_entry = marks_by_student.get(student.pk)
        aim, late, total = _compile_values_from_mark(mark_entry, aim_f, late_f)
        enrollment = str(student.enrollment_no or '').strip()
        name = (student.name or '').strip().upper()
        values = [
            (1, idx, ALIGN_C, None),
            (2, student.batch or '', ALIGN_C, None),
            (3, enrollment, Alignment(horizontal='center', vertical='center', wrap_text=True), '@'),
            (4, name, Alignment(horizontal='left', vertical='top', wrap_text=True), None),
            (5, str(student.roll_no or ''), ALIGN_C, None),
            (6, aim, ALIGN_C, None),
            (7, late, ALIGN_C, None),
            (8, total, ALIGN_C, None),
        ]
        for col, val, align, num_fmt in values:
            cell = ws.cell(row, col, val if val != '' else None)
            _style(cell, alignment=align)
            if num_fmt:
                cell.number_format = num_fmt
        row += 1

    data_end = max(row - 1, data_start)
    _apply_compiled_row_heights(ws, header_row, data_start, data_end, last_col)
    _apply_compiled_print_setup(ws, last_col, data_end)


def generate_compiled_marksheet_workbook(
    exam_type,
    subject,
    department,
    semester_label,
    department_label,
    template=None,
):
    """COMPILE marksheet — all batches on one sheet with aim, late, and total only."""
    schema = (template.schema if template else None) or {}
    aim_f, late_f = _compile_field_names(schema)
    max_marks = subject.max_marks_ipe if exam_type == 'IPE' else subject.max_marks_gp

    students_qs = Student.objects.filter(department=department)
    batches = list(students_qs.values_list('batch', flat=True).distinct().order_by('batch'))
    all_students = _sort_students_by_batch_and_roll(students_qs)
    marks_by_student = {}
    for batch in batches:
        marks_by_student.update(
            get_marks_by_batch(department, subject, exam_type, batch),
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'COMPILE'
    _build_compiled_sheet(
        ws, all_students, marks_by_student, subject, exam_type,
        semester_label, department_label, aim_f, late_f, max_marks,
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = (subject.name or subject.code or 'Subject').replace(' ', '_')
    safe_dept = department.name.replace(' ', '_')
    filename = f'Compile_Marksheet_{exam_type}_{safe_name}_{safe_dept}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _parse_mark_numeric(raw):
    if raw in (None, ''):
        return None
    if str(raw).strip().upper() == 'AB':
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def compute_mark_entry_total(editable_cols, mark_data):
    """
    Total = sum of Aim + Late + program marks (enter deductions as -2, -1, etc.).
    Final chit and remarks are excluded.
    """
    total = 0
    has_numeric = False
    for col in editable_cols:
        ctype = col['type']
        if ctype in ('remarks', 'total', 'final_chit'):
            continue
        fname = col.get('field_name', col['key'])
        n = _parse_mark_numeric(mark_data.get(fname, ''))
        if n is None:
            continue
        if ctype in ('mark', 'deduction_aim', 'deduction_late'):
            total += n
            has_numeric = True
    return total, has_numeric


def _editable_cols_for_duty(duty):
    template = resolve_marksheet_template(duty.exam_type, duty.subject, duty.department)
    if not template:
        return None
    schema = template.schema or {}
    return [
        c for c in schema.get('columns', [])
        if c['type'] not in ('sr_no', 'batch', 'enrollment_no', 'name', 'roll_no', 'total')
    ]


def is_student_duty_mark_complete(mark, editable_cols):
    """True when a student row has all required marks saved for this duty."""
    if mark is None:
        return False
    data = mark.mark_data or {}
    required = [c for c in editable_cols if c['type'] in ('mark', 'final_chit')]
    if not required:
        return bool(data) or bool((mark.remarks or '').strip())
    for col in required:
        fname = col.get('field_name', col['key'])
        if not str(data.get(fname, '')).strip():
            return False
    return True


def get_duty_marks_status(duty):
    """
    Return 'completed' when every student in the duty batch has required marks saved,
    otherwise 'pending'.
    """
    students = list(Student.objects.filter(department=duty.department, batch=duty.batch))
    if not students:
        return 'pending'
    editable_cols = _editable_cols_for_duty(duty)
    if not editable_cols:
        return 'pending'
    student_ids = [s.pk for s in students]
    existing = {
        m.student_id: m
        for m in MarkEntry.objects.filter(duty_assignment=duty, student_id__in=student_ids)
    }
    for stu in students:
        if not is_student_duty_mark_complete(existing.get(stu.pk), editable_cols):
            return 'pending'
    return 'completed'


def build_duty_marksheet_page_context(duty):
    """Build template context for faculty/admin duty marksheet grid."""
    students = sort_students_by_roll(Student.objects.filter(
        department=duty.department, batch=duty.batch,
    ))
    template = resolve_marksheet_template(duty.exam_type, duty.subject, duty.department)
    schema = template.schema if template else {}
    editable_cols = [
        c for c in schema.get('columns', [])
        if c['type'] not in ('sr_no', 'batch', 'enrollment_no', 'name', 'roll_no', 'total')
    ]
    existing = {
        m.student_id: m
        for m in MarkEntry.objects.filter(duty_assignment=duty, student__in=students)
    }
    student_rows = []
    for s in students:
        mark = existing.get(s.pk)
        data = (mark.mark_data if mark else {}) or {}
        cells = []
        for col in editable_cols:
            if col['type'] == 'remarks':
                val = (mark.remarks if mark else '') or data.get('remarks', '')
            else:
                fname = col.get('field_name', col['key'])
                val = data.get(fname, '')
            cells.append({'col': col, 'value': val})
        total_val, has_total = compute_mark_entry_total(editable_cols, data)
        if not has_total:
            total_val = ''
        student_rows.append({'student': s, 'mark': mark, 'cells': cells, 'total': total_val})

    column_layout = build_mark_entry_layout(editable_cols)
    return {
        'duty': duty,
        'student_rows': student_rows,
        'editable_cols': editable_cols,
        'column_layout': column_layout,
        'has_mark_groups': any(b['kind'] == 'group' for b in column_layout),
        'template': template,
        'marks_status': get_duty_marks_status(duty),
    }


def can_edit_duty_marks(user, duty):
    """Faculty and dept admin cannot edit when marks are locked; semester/super admin can."""
    if not duty.marks_locked:
        return True
    if user.is_superuser or user.role == User.Role.SUPER_ADMIN:
        return True
    if user.role == User.Role.SEMESTER_ADMIN:
        return True
    return False


def verify_and_lock_duty_marks(duty, user):
    """Mark duty marks as verified and locked by faculty."""
    duty.marks_locked = True
    duty.marks_locked_at = timezone.now()
    duty.marks_locked_by = user
    duty.save(update_fields=['marks_locked', 'marks_locked_at', 'marks_locked_by'])


def save_duty_marks(duty, post, students, user):
    """Persist faculty mark entry from POST using marksheet column fields."""
    template = resolve_marksheet_template(duty.exam_type, duty.subject, duty.department)
    schema = template.schema if template else {}
    editable = [
        c for c in schema.get('columns', [])
        if c['type'] not in ('sr_no', 'batch', 'enrollment_no', 'name', 'roll_no', 'total')
    ]

    for stu in students:
        mark_data = {}
        for col in editable:
            fname = col.get('field_name', col['key'])
            if col['type'] == 'remarks':
                continue
            val = post.get(f'{fname}_{stu.pk}', '').strip()
            if val:
                mark_data[fname] = val
        remarks = post.get(f'remarks_{stu.pk}', '').strip()

        numeric_sum, has_numeric = compute_mark_entry_total(editable, mark_data)

        if not mark_data and not remarks:
            MarkEntry.objects.filter(student=stu, duty_assignment=duty).delete()
            continue

        if has_numeric:
            mark_data['total'] = numeric_sum

        MarkEntry.objects.update_or_create(
            student=stu,
            duty_assignment=duty,
            defaults={
                'mark_data': mark_data,
                'marks_obtained': numeric_sum if has_numeric else 0,
                'remarks': remarks,
                'entered_by': user,
            },
        )

    now = timezone.now()
    FacultyDutyAssignment.objects.filter(pk=duty.pk).update(marks_saved_at=now)
    duty.marks_saved_at = now


def _short_program_title(label):
    """Shorten program group header for web UI."""
    text = _norm(label)
    m = re.search(r'program[\s-]*(\d+)', text, re.I)
    num = m.group(1) if m else ''
    max_m = _extract_max_marks(text)
    if num and max_m:
        return f'Program {num} ({max_m})'
    if num:
        return f'Program {num}'
    return text[:28] if text else 'Marks'


def _short_column_header(col):
    """Compact header label for a single marksheet column."""
    ctype = col['type']
    if ctype == 'final_chit':
        return {'short': 'Final Chit', 'sub': ''}
    if ctype == 'deduction_aim':
        return {'short': 'Aim Change', 'sub': '−2'}
    if ctype == 'deduction_late':
        return {'short': 'Late Coming', 'sub': '−1'}
    if ctype == 'remarks':
        return {'short': 'Remarks', 'sub': ''}
    if ctype == 'mark':
        sub = _norm(col.get('label9') or col.get('label8') or 'Marks')
        max_m = col.get('max_marks')
        short = sub
        if max_m and f'({max_m})' not in short:
            short = f'{sub} ({max_m})'
        return {'short': short, 'sub': ''}
    label = _norm(col.get('label9') or col.get('label8') or '')
    return {'short': label[:24], 'sub': ''}


def build_mark_entry_layout(editable_cols):
    """
    Ordered layout for mark-entry table headers.
    Returns list of {kind: 'single', col, header} or {kind: 'group', title, cols, headers}.
    """
    layout = []
    i = 0
    cols = list(editable_cols)
    while i < len(cols):
        col = cols[i]
        if col['type'] == 'mark':
            parent = col.get('parent8') or col.get('label8') or 'Program'
            group_cols = [col]
            i += 1
            while i < len(cols) and cols[i]['type'] == 'mark':
                nxt = cols[i]
                nxt_parent = nxt.get('parent8') or nxt.get('label8') or 'Program'
                if nxt_parent != parent:
                    break
                group_cols.append(nxt)
                i += 1
            layout.append({
                'kind': 'group',
                'title': _short_program_title(parent),
                'cols': group_cols,
                'headers': [_short_column_header(c) for c in group_cols],
            })
        else:
            layout.append({
                'kind': 'single',
                'col': col,
                'header': _short_column_header(col),
            })
            i += 1
    return layout


def ensure_default_marksheet_template():
    """Copy bundled IPE marksheet into media if missing."""
    from django.conf import settings

    dest_dir = Path(settings.BASE_DIR) / 'media' / 'marksheet_templates'
    dest_dir.mkdir(parents=True, exist_ok=True)
    source = Path(settings.BASE_DIR) / 'MARKSHEET OF IPE 2026.xlsx'
    if source.exists():
        dest = dest_dir / 'default_ipe_reference.xlsx'
        if not dest.exists():
            dest.write_bytes(source.read_bytes())
