import io
import math
from pathlib import Path

import openpyxl
from django.conf import settings
from django.http import HttpResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import FormField, GPGroup, GPGroupMemberDetail, Student

IPE_TITLE = 'Internal Practical Exam Attendance Sheet'
GP_TITLE = 'GP Attendance Sheet'

FONT_NAME = 'Times New Roman'
FONT_SIZE = 12
GP_FONT_SIZE = 10
IPE_LAST_COL = 11  # A through K
GP_GROUPS_PER_SHEET = 6

FONT_GP_NORMAL = Font(name=FONT_NAME, size=GP_FONT_SIZE, bold=False)
FONT_GP_BOLD = Font(name=FONT_NAME, size=GP_FONT_SIZE, bold=True)
FONT_GP_NOTES = Font(name=FONT_NAME, size=GP_FONT_SIZE, bold=True, color='FF0000')

THIN_SIDE = Side(style='thin', color='000000')
THICK_SIDE = Side(style='medium', color='000000')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# IPE palette
FILL_PEACH = PatternFill('solid', fgColor='FBD4B4')
FILL_GREY = PatternFill('solid', fgColor='D8D8D8')
FILL_TBL_HEAD = PatternFill('solid', fgColor='FDE9D9')
FILL_PURPLE = PatternFill('solid', fgColor='E5DFEC')

# GP palette (matches official GP attendance template)
GP_FILL_PEACH = PatternFill('solid', fgColor='F7CAAC')
GP_FILL_TBL_HEAD = PatternFill('solid', fgColor='E2EFD9')
GP_FILL_DIV = PatternFill('solid', fgColor='FFF2CC')
GP_FILL_GROUP_ALT = PatternFill('solid', fgColor='DDEBF7')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

FONT_BOLD = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
FONT_NORMAL = Font(name=FONT_NAME, size=FONT_SIZE, bold=False)
FONT_NOTES = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='FF0000')


def get_default_template_path(exam_type):
    base = Path(settings.BASE_DIR) / 'media' / 'attendance_templates'
    filename = 'ipe_template.xlsx' if exam_type == 'IPE' else 'gp_template.xlsx'
    path = base / filename
    if path.exists():
        return str(path)
    fallback = base / 'ipe_template.xlsx'
    return str(fallback) if fallback.exists() else None


def _style_cell(cell, font=None, fill=None, alignment=None, border=True, border_obj=None):
    cell.font = font or FONT_NORMAL
    if fill:
        cell.fill = fill
    cell.alignment = alignment or ALIGN_LEFT
    if border_obj:
        cell.border = border_obj
    elif border:
        cell.border = THIN_BORDER


def _merge_and_style(ws, range_str, value, font=None, fill=None, alignment=None):
    ws.merge_cells(range_str)
    cell = ws[range_str.split(':')[0]]
    cell.value = value
    fnt = font or FONT_BOLD
    aln = alignment or ALIGN_CENTER
    start, end = range_str.split(':')
    start_col = openpyxl.utils.column_index_from_string(''.join(c for c in start if c.isalpha()))
    start_row = int(''.join(c for c in start if c.isdigit()))
    end_col = openpyxl.utils.column_index_from_string(''.join(c for c in end if c.isalpha()))
    end_row = int(''.join(c for c in end if c.isdigit()))
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            _style_cell(ws.cell(r, c), font=fnt, fill=fill, alignment=aln)


def _set_ipe_column_widths(ws):
    widths = {'A': 6, 'B': 8, 'C': 18, 'D': 14, 'E': 14, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# Landscape A4 GP attendance — proportional to official LJIET template (wide name / sign cols).
GP_LANDSCAPE_WIDTHS = {
    1: 7,    # Sr No
    2: 6,    # Div
    3: 12,   # Group ID
    4: 9,    # Roll No
    5: 18,   # Enrollment
    6: 38,   # Name of Student
}


def _gp_column_width_limits(col, last_col, sign_col):
    """Min/max Excel column width for landscape GP attendance."""
    limits = {
        1: (6, 8),
        2: (5, 7),
        3: (10, 14),
        4: (7, 10),
        5: (15, 20),
        6: (30, 45),
    }
    if col == sign_col:
        return (18, 32)
    if col in limits:
        return limits[col]
    return (10, 16)


def _gp_chars_per_line(col_width, font_size=GP_FONT_SIZE):
    """Approximate visible characters per line for Times New Roman at given column width."""
    factor = 0.85 if font_size <= 10 else 0.75
    return max(6, int(col_width * factor))


def _gp_row_height_for_text(text, col_width, font_size=GP_FONT_SIZE, min_height=14, line_height=13):
    """Row height (points) from wrapped text in a column."""
    if not text:
        return min_height
    text = str(text).strip()
    if not text:
        return min_height
    chars = _gp_chars_per_line(col_width, font_size)
    lines = max(1, math.ceil(len(text) / chars))
    return min(min_height + (lines - 1) * line_height, 48)


def _gp_display_name(student):
    """Format student name for attendance sheet (full name, uppercase like official template)."""
    name = (student.name or '').strip()
    return name.upper() if name else ''


def _gp_cell_str(value):
    """Render cell value as plain string (avoids number formatting on enrollment/roll)."""
    if value is None:
        return ''
    return str(value).strip()


def _autofit_gp_columns(ws, last_col, sign_col, header_row, data_start, data_end):
    """Set landscape-friendly column widths from content and official template proportions."""
    for col in range(1, last_col + 1):
        min_w, max_w = _gp_column_width_limits(col, last_col, sign_col)
        base = GP_LANDSCAPE_WIDTHS.get(col)
        max_len = 0
        header_val = ws.cell(header_row, col).value
        if header_val:
            for line in str(header_val).split('\n'):
                max_len = max(max_len, len(line))
        for row in range(data_start, data_end + 1):
            val = ws.cell(row, col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        if base is not None:
            width = min(max(max_len + 2, base, min_w), max_w)
        elif col == sign_col:
            width = min(max(max_len + 2, 22, min_w), max_w)
        else:
            width = min(max(max_len + 1, min_w), max_w)
        ws.column_dimensions[get_column_letter(col)].width = width


def _autofit_gp_row_heights(ws, name_col, header_row, data_start, data_end, notes_row=7):
    """Dynamically set row heights from wrapped content (names, headers, notes)."""
    name_width = ws.column_dimensions[get_column_letter(name_col)].width or 22

    notes = ws.cell(notes_row, 1).value
    if notes:
        note_lines = str(notes).count('\n') + 1
        ws.row_dimensions[notes_row].height = min(14 + note_lines * 12, 52)

    header_lines = 1
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val:
            header_lines = max(header_lines, str(val).count('\n') + 1)
    ws.row_dimensions[header_row].height = min(14 + header_lines * 11, 36)

    for row in range(data_start, data_end + 1):
        name_val = ws.cell(row, name_col).value
        if not name_val:
            ws.row_dimensions[row].height = 15
            continue
        col_width = name_width
        chars_per_line = _gp_chars_per_line(col_width)
        lines_needed = math.ceil(len(str(name_val)) / chars_per_line)
        wrap = lines_needed > 1
        cell = ws.cell(row, name_col)
        cell.alignment = Alignment(
            horizontal='left', vertical='center',
            wrap_text=wrap, shrink_to_fit=False,
        )
        ws.row_dimensions[row].height = (
            _gp_row_height_for_text(name_val, col_width) if wrap else 15
        )


def _apply_print_setup(ws, last_col, last_row, landscape=True):
    """A4 print — fit entire sheet to one page for hard-copy printing."""
    from openpyxl.worksheet.page import PageMargins

    last_letter = get_column_letter(last_col)
    ws.print_area = f'A1:{last_letter}{last_row}'
    ws.page_setup.orientation = 'landscape' if landscape else 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None
    ws.page_margins = PageMargins(
        left=0.2, right=0.2, top=0.25, bottom=0.25, header=0.1, footer=0.1,
    )
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False


def _apply_gp_print_setup(ws, last_col, last_row):
    _apply_print_setup(ws, last_col, last_row, landscape=True)


def _build_ipe_sheet(ws, students, batch_name, semester_label, department_label,
                     subject_name, subject_code):
    """Build IPE attendance sheet — one sheet per batch, all students listed."""
    ws.sheet_view.showGridLines = True
    _set_ipe_column_widths(ws)

    _merge_and_style(ws, 'A1:K1', 'L. J. University', fill=FILL_PEACH)
    _merge_and_style(ws, 'A2:K2', 'L. J. Institute of Engineering & Technology, Ahmedabad', fill=FILL_PEACH)
    _merge_and_style(
        ws, 'A3:K3',
        f'Semester - {semester_label}   {department_label} Department',
        fill=FILL_PEACH,
    )
    _merge_and_style(ws, 'A4:K4', IPE_TITLE, fill=FILL_GREY)

    _merge_and_style(ws, 'A5:D6', f'SUBJECT NAME: {subject_name}', alignment=ALIGN_LEFT)
    _merge_and_style(ws, 'E5:K6', 'Date:', alignment=ALIGN_LEFT)
    _merge_and_style(ws, 'A7:D8', f'SUBJECTCODE: {subject_code}', alignment=ALIGN_LEFT)
    _merge_and_style(ws, 'E7:G8', 'START TIME:', alignment=ALIGN_LEFT)
    _merge_and_style(ws, 'H7:K8', 'END TIME:', alignment=ALIGN_LEFT)

    headers = [
        ('A9:A10', 'Sr.\nNo.', FILL_TBL_HEAD),
        ('B9:B10', 'Batch', FILL_TBL_HEAD),
        ('C9:C10', 'Enrollment Number', FILL_TBL_HEAD),
        ('D9:G10', 'Name of Student', FILL_TBL_HEAD),
        ('H9:H10', 'Roll No', FILL_TBL_HEAD),
        ('I9:J10', 'Sign', FILL_PURPLE),
        ('K9:K10', 'Remarks', FILL_PURPLE),
    ]
    for merge_range, text, fill in headers:
        _merge_and_style(ws, merge_range, text, fill=fill)

    data_start = 11
    student_count = len(students)
    if student_count > 32:
        data_row_height = 12
        header_block_height = 18
    elif student_count > 24:
        data_row_height = 13
        header_block_height = 20
    else:
        data_row_height = 15
        header_block_height = 22

    for idx, stu in enumerate(students, start=1):
        row = data_start + idx - 1
        ws.merge_cells(f'D{row}:G{row}')
        ws.merge_cells(f'I{row}:J{row}')

        _style_cell(ws.cell(row, 1, idx), alignment=ALIGN_CENTER)
        _style_cell(ws.cell(row, 2, batch_name), alignment=ALIGN_CENTER)
        _style_cell(ws.cell(row, 3, stu.enrollment_no), alignment=ALIGN_CENTER)
        _style_cell(ws.cell(row, 4, stu.name), alignment=ALIGN_LEFT)
        for c in range(5, 8):
            _style_cell(ws.cell(row, c), alignment=ALIGN_LEFT)
        _style_cell(ws.cell(row, 8, stu.roll_no), alignment=ALIGN_CENTER)
        for c in range(9, 12):
            _style_cell(ws.cell(row, c), alignment=ALIGN_CENTER)
        ws.row_dimensions[row].height = data_row_height

    footer_start = data_start + student_count
    footers = [
        'NAME OF EXTERNAL EXAMINER: ',
        'SIGNATURE OF EXTERNAL EXAMINER:',
        'NAME OF INTERNAL EXAMINER:                     ',
        'SIGNATURE OF INTERNAL EXAMINER:',
    ]
    footer_end = footer_start + len(footers) - 1
    for i, text in enumerate(footers):
        row = footer_start + i
        _merge_and_style(ws, f'A{row}:K{row}', text, alignment=ALIGN_LEFT)
        ws.row_dimensions[row].height = 14

    for r in range(1, 5):
        ws.row_dimensions[r].height = header_block_height
    ws.row_dimensions[9].height = 26
    ws.row_dimensions[10].height = 26

    _apply_print_setup(ws, IPE_LAST_COL, footer_end, landscape=True)


def _gp_extra_columns(template):
    """Dynamic columns from GP form template (excluding built-in attendance fields)."""
    if not template:
        return []
    from .gp_utils import get_group_fields, get_member_fields

    skip = ('group id', 'case', 'project title', 'gender diversity', 'religion diversity',
            'enrollment', 'roll', 'div', 'gender', 'faculty')
    extras = []
    seen = set()
    for field in list(get_group_fields(template)) + list(get_member_fields(template)):
        low = field.field_label.lower()
        if any(s in low for s in skip):
            continue
        if low in seen:
            continue
        seen.add(low)
        extras.append(field)
    return extras


def _build_gp_attendance_sheet(ws, groups, batch_name, group_start_num,
                               semester_label, department_label,
                               subject_name, subject_code, template=None):
    """
    Build one GP attendance sheet for up to GP_GROUPS_PER_SHEET groups.
    Layout matches the official LJIET GP attendance template.
    """
    from .gp_utils import _batch_palette, _read_group_id

    extra_fields = _gp_extra_columns(template)
    sign_col = 7 + len(extra_fields)
    last_col = sign_col
    last_letter = get_column_letter(last_col)
    sign_letter = get_column_letter(sign_col)

    ws.sheet_view.showGridLines = True

    gp_hdr_font = FONT_GP_BOLD
    header_range = f'A1:{last_letter}1'
    _merge_and_style(ws, header_range, 'L. J. University', font=gp_hdr_font, fill=GP_FILL_PEACH)
    _merge_and_style(
        ws, f'A2:{last_letter}2',
        'L. J. Institute of Engineering & Technology, Ahmedabad',
        font=gp_hdr_font, fill=GP_FILL_PEACH,
    )
    _merge_and_style(
        ws, f'A3:{last_letter}3',
        f'Semester - {semester_label}   {department_label} Department',
        font=gp_hdr_font, fill=GP_FILL_PEACH,
    )
    _merge_and_style(ws, f'A4:{last_letter}4', GP_TITLE, font=gp_hdr_font, fill=FILL_GREY)

    date_col = get_column_letter(max(sign_col - 1, 7))
    _merge_and_style(
        ws, f'A5:{get_column_letter(sign_col - 1)}5',
        f'SUBJECT NAME: {subject_name}',
        font=FONT_GP_BOLD, alignment=ALIGN_LEFT,
    )
    _merge_and_style(ws, f'{date_col}5:{sign_letter}5', 'Date:', font=FONT_GP_BOLD, alignment=ALIGN_LEFT)
    _merge_and_style(
        ws, f'A6:{get_column_letter(sign_col - 1)}6',
        f'SUBJECTCODE: {subject_code}',
        font=FONT_GP_BOLD, alignment=ALIGN_LEFT,
    )
    _merge_and_style(ws, f'{date_col}6:{sign_letter}6', 'Time:', font=FONT_GP_BOLD, alignment=ALIGN_LEFT)

    notes = (
        'N.B : \n'
        '1) For absent students, enter "AB" in each marks column.\n'
        '2) Do not merge any entries.'
    )
    _merge_and_style(ws, f'A7:{sign_letter}7', notes, font=FONT_GP_NOTES, alignment=ALIGN_LEFT)

    std_headers = [
        (1, 'Sr \nNo.'),
        (2, 'Div'),
        (3, 'Group ID'),
        (4, 'Roll No.'),
        (5, 'Enrollment No'),
        (6, 'Name of Student'),
    ]
    name_col = 6
    for col, label in std_headers:
        cell = ws.cell(row=8, column=col, value=label)
        _style_cell(cell, font=FONT_GP_BOLD, fill=GP_FILL_TBL_HEAD, alignment=ALIGN_CENTER)
    for i, field in enumerate(extra_fields):
        col = 7 + i
        cell = ws.cell(row=8, column=col, value=field.field_label)
        _style_cell(cell, font=FONT_GP_BOLD, fill=GP_FILL_TBL_HEAD, alignment=ALIGN_CENTER)
    cell = ws.cell(row=8, column=sign_col, value='SIGN')
    _style_cell(cell, font=FONT_GP_BOLD, fill=GP_FILL_TBL_HEAD, alignment=ALIGN_CENTER)

    all_batches = {batch_name}
    light, dark = _batch_palette(batch_name, all_batches)
    data_start = 9
    data_row = data_start
    sr_no = 1
    merge_ranges = []
    group_end_rows = []
    group_idx = 0

    for group in groups:
        group_idx += 1
        details = list(group.member_details.select_related('student').order_by('student__roll_no'))
        if not details:
            for m in group.members.order_by('roll_no'):
                details.append(GPGroupMemberDetail(student=m, gender='', member_data={}))
        if not details:
            continue

        gid = _read_group_id(group) or f'{batch_name}_{group_start_num + group_idx - 1}'
        start_row = data_row
        group_fill = PatternFill('solid', fgColor=dark if group_idx % 2 == 0 else light)

        for mem_idx, detail in enumerate(details):
            s = detail.student
            member_data = getattr(detail, 'member_data', {}) or {}
            is_last = mem_idx == len(details) - 1
            row_border = Border(
                left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE,
                bottom=THICK_SIDE if is_last else THIN_SIDE,
            )

            display_name = _gp_display_name(s)
            row_vals = {
                1: sr_no,
                2: batch_name if mem_idx == 0 and data_row == data_start else None,
                3: gid if mem_idx == 0 else None,
                4: _gp_cell_str(s.roll_no),
                5: _gp_cell_str(s.enrollment_no),
                6: display_name,
            }
            for i, field in enumerate(extra_fields):
                col = 7 + i
                if field.field_scope == FormField.FieldScope.GROUP:
                    row_vals[col] = group.group_data.get(field.field_name, '') if mem_idx == 0 else ''
                else:
                    row_vals[col] = member_data.get(field.field_name, '')

            for col in range(1, last_col + 1):
                val = row_vals.get(col, '')
                cell = ws.cell(row=data_row, column=col, value=val)
                fill = group_fill
                if col == 2 and val:
                    fill = GP_FILL_DIV
                elif col == 3 and val:
                    fill = GP_FILL_GROUP_ALT
                align = ALIGN_LEFT if col == name_col else ALIGN_CENTER
                _style_cell(
                    cell, font=FONT_GP_NORMAL, fill=fill,
                    alignment=align, border_obj=row_border,
                )

            sr_no += 1
            data_row += 1

        end_row = data_row - 1
        if end_row >= start_row:
            merge_ranges.append((start_row, end_row, 3))
            group_end_rows.append(end_row)

    data_end = data_row - 1
    if data_end >= data_start:
        ws.merge_cells(f'B{data_start}:B{data_end}')
        div_cell = ws.cell(data_start, 2, batch_name)
        _style_cell(div_cell, fill=GP_FILL_DIV, alignment=ALIGN_CENTER)

    for start_row, end_row, col in merge_ranges:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        merged = ws.cell(row=start_row, column=col)
        merged.fill = GP_FILL_GROUP_ALT
        merged.alignment = ALIGN_CENTER

    for end_row in group_end_rows:
        for col in range(1, last_col + 1):
            cell = ws.cell(row=end_row, column=col)
            cell.border = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THICK_SIDE)

    footer_start = data_end + 1 if data_end >= data_start else data_start
    footers = [
        'NAME OF EXTERNAL EXAMINER: ',
        'SIGNATURE OF EXTERNAL EXAMINER:',
        'NAME OF INTERNAL EXAMINER:',
        'SIGNATURE OF INTERNAL EXAMINER:',
    ]
    footer_end = footer_start + len(footers) - 1
    for i, text in enumerate(footers):
        row = footer_start + i
        _merge_and_style(ws, f'A{row}:{sign_letter}{row}', text, font=FONT_GP_NORMAL, alignment=ALIGN_LEFT)
        ws.row_dimensions[row].height = 16

    for r in range(1, 5):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 16

    if data_end >= data_start:
        _autofit_gp_columns(ws, last_col, sign_col, header_row=8, data_start=data_start, data_end=data_end)
        _autofit_gp_row_heights(ws, name_col, header_row=8, data_start=data_start, data_end=data_end, notes_row=7)

    _apply_gp_print_setup(ws, last_col, footer_end)


def _gp_sheet_title(batch, start_num, end_num):
    return f'{batch}({start_num} TO {end_num}) GP'[:31]


def generate_gp_attendance_workbook(department, subject, semester_label, department_label, batch_filter=None):
    """Generate GP attendance workbook — multiple sheets per batch (6 groups per sheet)."""
    from .gp_utils import get_gp_template, _group_sort_key

    template = get_gp_template(department)
    groups = list(
        GPGroup.objects.filter(
            department=department,
            subject=subject,
            is_submitted=True,
        ).select_related('leader', 'project_case').prefetch_related('member_details__student', 'members')
    )
    groups.sort(key=_group_sort_key)

    if batch_filter:
        groups = [g for g in groups if g.leader and g.leader.batch == batch_filter]

    batches = sorted({g.leader.batch for g in groups if g.leader and g.leader.batch})
    if not batches:
        batches = ['A1']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for batch in batches:
        batch_groups = [g for g in groups if g.leader and g.leader.batch == batch]
        if not batch_groups:
            continue

        chunks = [
            batch_groups[i:i + GP_GROUPS_PER_SHEET]
            for i in range(0, len(batch_groups), GP_GROUPS_PER_SHEET)
        ]
        for chunk_idx, chunk in enumerate(chunks):
            start_num = chunk_idx * GP_GROUPS_PER_SHEET + 1
            end_num = start_num + len(chunk) - 1
            ws = wb.create_sheet(title=_gp_sheet_title(batch, start_num, end_num))
            _build_gp_attendance_sheet(
                ws, chunk, batch, start_num,
                semester_label, department_label,
                subject.name, subject.code or '',
                template=template,
            )

    if not wb.sheetnames:
        ws = wb.create_sheet(title='No Data')
        ws.cell(1, 1, 'No GP submissions found for this subject.')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_subject = (subject.code or subject.name).replace(' ', '_')
    batch_suffix = f'_{batch_filter}' if batch_filter else ''
    filename = f'GP_Attendance_{safe_subject}_{department.name}{batch_suffix}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_ipe_attendance_workbook(department, subject, semester_label, department_label):
    """Generate IPE attendance workbook — one sheet per batch."""
    students_qs = Student.objects.filter(department=department).order_by('batch', 'roll_no')
    batches = list(students_qs.values_list('batch', flat=True).distinct().order_by('batch'))
    if not batches:
        batches = ['A1']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for batch in batches:
        batch_students = list(students_qs.filter(batch=batch))
        ws = wb.create_sheet(title=str(batch)[:31])
        _build_ipe_sheet(
            ws, batch_students, batch,
            semester_label, department_label,
            subject.name, subject.code or '',
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_subject = subject.code or subject.name.replace(' ', '_')
    filename = f'Attendance_IPE_{safe_subject}_{department.name}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_attendance_workbook(
    template_path,
    department,
    subject,
    exam_type,
    semester_label,
    department_label,
    batch_filter=None,
):
    """Generate attendance workbook for IPE or GP."""
    if exam_type == 'GP':
        return generate_gp_attendance_workbook(
            department, subject, semester_label, department_label, batch_filter=batch_filter,
        )
    return generate_ipe_attendance_workbook(
        department, subject, semester_label, department_label,
    )


def ensure_default_templates():
    """Copy bundled templates into media if missing."""
    dest_dir = Path(settings.BASE_DIR) / 'media' / 'attendance_templates'
    dest_dir.mkdir(parents=True, exist_ok=True)

    ipe_source = Path(settings.BASE_DIR) / 'tgh thfdghg.xlsx'
    if ipe_source.exists():
        dest = dest_dir / 'ipe_template.xlsx'
        if not dest.exists():
            dest.write_bytes(ipe_source.read_bytes())

    gp_source = Path(settings.BASE_DIR) / 'Attedance sheet A1 A2 A3 A9.xlsx'
    if gp_source.exists():
        dest = dest_dir / 'gp_template.xlsx'
        if not dest.exists():
            dest.write_bytes(gp_source.read_bytes())
